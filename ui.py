#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EVARISIS Gestor HUV - Sistema de Oncología
Punto de entrada principal de la aplicación - Migrado completamente a TTKBootstrap

Este script se encarga de:
1. Configurar la ruta del ejecutable de Tesseract OCR.
2. Iniciar la interfaz gráfica de usuario moderna (dashboard).
"""

import ttkbootstrap as ttk
from ttkbootstrap import Style
from ttkbootstrap.constants import *
import tkinter.ttk as ttk_std
from PIL import Image, ImageTk
from tkinter import filedialog, messagebox
import tkinter as tk
import threading
import os
import re
import pandas as pd
import seaborn as sns
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from datetime import datetime
import numpy as np
import argparse
import sys
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import configparser
import pytesseract
from pathlib import Path

from calendario import CalendarioInteligente
from huv_web_automation import automatizar_entrega_resultados, Credenciales
from version_info import get_version_string, get_build_info, get_full_version_info, get_dependencies_actual

# =========================
# Configuración de Tesseract OCR
# =========================

def configure_tesseract():
    """
    Lee el archivo config.ini para encontrar la ruta de Tesseract OCR
    y la configura para que Pytesseract pueda utilizarla.
    """
    try:
        config = configparser.ConfigParser(interpolation=None)
        config_path = Path(__file__).resolve().parent / 'config.ini'
        config.read(config_path, encoding='utf-8')

        tesseract_cmd = None
        if sys.platform.startswith("win"):
            tesseract_cmd = config.get('PATHS', 'WINDOWS_TESSERACT', fallback=None)
        elif sys.platform.startswith("darwin"):
            tesseract_cmd = config.get('PATHS', 'MACOS_TESSERACT', fallback=None)
        else: # Asumimos Linux/otro
            tesseract_cmd = config.get('PATHS', 'LINUX_TESSERACT', fallback=None)

        if tesseract_cmd and Path(tesseract_cmd).exists():
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
            print(f"Tesseract OCR configurado en: {tesseract_cmd}")
        else:
            print("ADVERTENCIA: No se encontró la ruta de Tesseract en config.ini o la ruta no es válida.")
            print("El sistema intentará usar la variable de entorno PATH.")

    except Exception as e:
        print(f"Error al configurar Tesseract desde config.ini: {e}")
        print("Se continuará usando la configuración por defecto de Pytesseract.")

# =========================
# Configuración de temas TTKBootstrap
# =========================

# Mapeo de argumentos de tema a temas TTKBootstrap
THEME_MAP = {
    "dark": "darkly",
    "light": "flatly", 
    "blue": "cosmo",
    "professional": "litera",
    "medical": "pulse",
    "modern": "superhero",
    "classic": "journal"
}

# Paleta de colores base (se ajustará según el tema)
COLORS = {
    "accent": "#2b6cb0",
    "bg": "#ffffff", 
    "surface": "#f8f9fa",
    "text": "#212529",
    "muted": "#6c757d"
}

# Estilo visual para seaborn/mpl dentro de Tk
sns.set_theme(
    style="darkgrid",
    rc={
        "axes.facecolor": "#343638",
        "grid.color": "#4a4d50",
        "figure.facecolor": "#2b2b2b",
        "text.color": "white",
        "xtick.color": "white",
        "ytick.color": "white",
        "axes.labelcolor": "white",
        "axes.titlecolor": "white",
    },
)

# Módulos del proyecto
import procesador_ihq_biomarcadores
import database_manager


class App(ttk.Window):
    def __init__(self, info_usuario=None, tema="superhero"):
        # Inicializar TTKBootstrap Window con el tema
        super().__init__(themename=tema)
        
        self.title("EVARISIS Gestor HUV - Oncología")
        self.state('zoomed')  # Maximizar ventana

        # Información del usuario
        self.info_usuario = info_usuario or {"nombre": "Invitado", "cargo": "N/A", "ruta_foto": "SIN_FOTO", "ruta_directorio_fotos": ""}
        
        # Fuentes estándar para consistencia - usando las mismas del proyecto estadístico
        self.FONT_TITULO = ("Segoe UI", 22, "bold")
        self.FONT_SUBTITULO = ("Segoe UI", 12)
        self.FONT_NORMAL = ("Segoe UI", 11)
        self.FONT_ETIQUETA = ("Segoe UI", 9, "italic")
        self.FONT_NOMBRE_PERFIL = ("Segoe UI", 16, "bold")
        self.FONT_BOTONES = ("Segoe UI", 12)
        
        # Configurar tema actual
        self.current_theme = tema
        self.temas_disponibles = [
            'superhero', 'flatly', 'cyborg', 'journal', 'solar', 'darkly', 
            'minty', 'pulse', 'sandstone', 'united', 'morph', 'vapor', 
            'yeti', 'cosmo', 'litera', 'lumen', 'simplex', 'zephyr'
        ]
        
        # Configurar iconos y foto del usuario
        self.iconos = self._cargar_iconos()
        self.foto_usuario = self._cargar_foto_usuario()

        # Estados del nuevo sistema de navegación flotante
        self.header_visible = True
        self.floating_menu_visible = False
        self.welcome_screen_active = True
        self.current_view = "welcome"  # welcome, database, visualizar, dashboard
        
        # Crear la interfaz
        self._create_layout()

    def _create_layout(self):
        """Crear la interfaz principal usando TTKBootstrap"""
        # Crear layout principal similar al proyecto estadístico
        main_frame = ttk.Frame(self, padding=0)
        main_frame.pack(expand=True, fill=BOTH)

        # ===== Header institucional =====
        self._create_header(main_frame)

        # Variables necesarias ANTES de crear la interfaz
        self.master_df = pd.DataFrame()  # DataFrame maestro (fuente única de verdad)
        self._compare_controls = {}
        
        # Inicializar componentes que serán referenciados (para evitar AttributeError)
        self.cmb_servicio = None
        self.cmb_malig = None
        self.cmb_resp = None
        self.tree = None

        # ===== Separador =====
        ttk.Separator(main_frame, orient=HORIZONTAL).pack(fill=X, padx=20, pady=5)

        # ===== Contenido principal (sin sidebar tradicional) =====
        self._create_main_content(main_frame)
        
        # ===== Menú flotante =====
        self._create_floating_menu()
        
        # ===== Botón flotante =====
        self._create_floating_button()

        # Inicializar estilo de treeview
        self._init_treeview_style()
        
        # Crear y mostrar pantalla de bienvenida inmediatamente
        self._create_welcome_screen()
        self.after(50, self.show_welcome_screen)

    def _create_header(self, parent):
        """Crear header institucional"""
        self.header = ttk.Frame(parent, padding=(20, 10))
        self.header.pack(fill=X)

        # Logo izquierdo
        left = ttk.Frame(self.header)
        left.pack(side=LEFT, padx=(0, 16))
        if self.iconos.get("logo1"):
            ttk.Label(left, image=self.iconos["logo1"]).pack()

        # Centro - Título y subtítulo
        center = ttk.Frame(self.header)
        center.pack(side=LEFT, expand=True)
        
        ttk.Label(
            center,
            text="EVARISIS Gestor HUV - Oncología",
            font=self.FONT_TITULO,
            anchor=W
        ).pack(fill=X)
        
        ttk.Label(
            center,
            text="Suite de procesamiento y análisis oncológico del Hospital Universitario del Valle",
            font=self.FONT_NORMAL,
            anchor=W,
            bootstyle=SECONDARY
        ).pack(fill=X, pady=(2, 0))

        # Perfil derecho
        right = ttk.Frame(self.header)
        right.pack(side=RIGHT)

        # Tarjeta de perfil profesional
        profile_card = ttk.Frame(right, padding=(10, 8), bootstyle="secondary")
        profile_card.pack(side=RIGHT, padx=(10, 0))
        
        # Foto del usuario si existe
        if self.foto_usuario:
            ttk.Label(profile_card, image=self.foto_usuario).pack(side=LEFT, padx=(0, 10))
        
        # Datos del usuario
        datos = ttk.Frame(profile_card)
        datos.pack(side=LEFT)
        ttk.Label(datos, text=self.info_usuario.get("nombre", "Invitado"), font=self.FONT_NOMBRE_PERFIL).pack(anchor=W)
        ttk.Label(datos, text=self.info_usuario.get("cargo", "N/A"), font=self.FONT_SUBTITULO, bootstyle=INFO).pack(anchor=W)
        
        # Botón de información de versión
        version_btn = ttk.Button(
            right,
            text=f"v{get_version_string().split('-')[0].replace('v', '')}",
            command=self._show_version_info,
            bootstyle="info-outline",
            width=8
        )
        version_btn.pack(side=RIGHT, padx=(10, 10))
        
        # Logo derecho
        if self.iconos.get("logo3"):
            ttk.Label(right, image=self.iconos["logo3"]).pack(side=RIGHT, padx=(10, 0))

    def _create_main_content(self, parent):
        """Crear el contenido principal sin sidebar tradicional"""
        # Contenedor principal que ocupa toda la ventana
        self.content_container = ttk.Frame(parent, padding=0)
        self.content_container.pack(expand=True, fill=BOTH)

        # Crear los diferentes paneles de contenido
        self._create_content_panels()

    def _create_floating_menu(self):
        """Crear el menú flotante con diseño profesional y sombras"""
        # Frame principal del menú con sombra
        self.floating_menu = ttk.Frame(
            self, 
            padding=5,
            relief="raised",
            borderwidth=3
        )
        self.floating_menu.place(x=-300, y=20, width=280, height=420)  # Inicialmente oculto, más arriba
        
        # Header del menú con gradiente visual
        header_frame = ttk.Frame(self.floating_menu, bootstyle="primary", padding=10)
        header_frame.pack(fill=X, pady=(0, 5))
        
        # Título con mejor diseño
        ttk.Label(
            header_frame,
            text="🏥 HUV ONCOLOGÍA",
            font=("Segoe UI", 13, "bold"),
            bootstyle="inverse-primary",
            anchor="center"
        ).pack(expand=True)
        
        ttk.Label(
            header_frame,
            text="Panel de Navegación",
            font=("Segoe UI", 9),
            bootstyle="inverse-primary",
            anchor="center"
        ).pack()
        
        # Separador elegante
        separator = ttk.Separator(self.floating_menu, orient="horizontal")
        separator.pack(fill=X, padx=10, pady=5)
        
        # Frame para botones de navegación
        nav_frame = ttk.Frame(self.floating_menu, padding=5)
        nav_frame.pack(fill=BOTH, expand=True)
        
        # Botones de navegación con mejor espaciado y diseño
        self.nav_buttons = {}
        nav_items = [
            ("🏠 Inicio", "home", "success-outline", self._nav_to_welcome),
            ("🗄️ Base de Datos", "database", "primary", self._nav_to_database),
            ("📊 Visualizar Datos", "informes", "info", self._nav_to_visualizar),
            ("📈 Dashboard", "dashboard", "warning", self._nav_to_dashboard),
            ("🔄 Procesador PDF", "web", "secondary", self._nav_to_web_auto),
            ("ℹ️ Acerca de", "about", "info-outline", self._show_version_info),
        ]
        
        for text, icon_key, style, callback in nav_items:
            # Frame para cada botón con efecto hover
            btn_frame = ttk.Frame(nav_frame, padding=2)
            btn_frame.pack(fill=X, pady=3)
            
            btn = ttk.Button(
                btn_frame,
                text=text,
                command=callback,
                bootstyle=style,
                width=28
            )
            btn.pack(fill=X)
            self.nav_buttons[text] = btn
            
            # Agregar efectos hover
            btn.bind("<Enter>", lambda e, b=btn: self._on_menu_btn_hover(b, True))
            btn.bind("<Leave>", lambda e, b=btn: self._on_menu_btn_hover(b, False))
        
        # Footer con botón de cerrar
        footer_frame = ttk.Frame(self.floating_menu, bootstyle="secondary", padding=5)
        footer_frame.pack(fill=X, side=BOTTOM, pady=(5, 0))
        
        close_btn = ttk.Button(
            footer_frame,
            text="✕ Cerrar Menú",
            command=self._toggle_floating_menu,
            bootstyle="danger-outline",
            width=25
        )
        close_btn.pack()
        
        # Variable para controlar estado del menú
        self.menu_is_open = False

    def _on_menu_btn_hover(self, button, entering):
        """Efectos hover para botones del menú"""
        if entering:
            # Efecto al entrar - escalar ligeramente
            button.configure(cursor="hand2")
            # Aquí se podría agregar más efectos visuales
        else:
            # Efecto al salir - restaurar
            button.configure(cursor="")

    def _create_floating_button(self):
        """Crear el botón flotante circular, compacto y natural"""
        # Frame contenedor circular más pequeño
        self.floating_btn_container = ttk.Frame(
            self,
            relief="raised",
            borderwidth=2
        )
        self.floating_btn_container.place(x=15, y=150, width=50, height=50)
        
        # Botón flotante circular y compacto
        self.floating_btn = ttk.Button(
            self.floating_btn_container,
            text="⚡",  # Icono más compacto y moderno
            command=self._toggle_floating_menu,
            bootstyle="info-outline",  # Estilo outline para apariencia más suave
            width=2,  # Más compacto
            cursor="hand2"
        )
        self.floating_btn.pack(expand=True, fill=BOTH, padx=3, pady=3)
        
        # Variables para animaciones con nueva posición
        self.floating_btn_base_y = 150
        self.floating_animation_running = False
        self.hover_animation_running = False
        
        # Configurar eventos de hover
        self.floating_btn.bind("<Enter>", self._on_btn_hover_enter)
        self.floating_btn.bind("<Leave>", self._on_btn_hover_leave)
        
        # Agregar tooltip visual
        self.floating_btn.bind("<Button-1>", lambda e: self.floating_btn.configure(text="✕" if not self.floating_menu_visible else "☰"))
        
        # Iniciar animación flotante continua
        self._start_floating_animation()

    def _start_floating_animation(self):
        """Iniciar animación flotante continua sutil"""
        if not self.floating_animation_running and not self.floating_menu_visible:
            self.floating_animation_running = True
            self._animate_floating(0)

    def _animate_floating(self, step):
        """Animación flotante sutil - movimiento vertical suave como flotando"""
        if self.floating_animation_running and not self.floating_menu_visible:
            import math
            # Movimiento sinusoidal más sutil para botón pequeño (±3 pixels)  
            offset = math.sin(step * 0.08) * 3
            new_y = self.floating_btn_base_y + offset
            
            # Actualizar posición del botón con dimensiones más pequeñas
            if hasattr(self, 'floating_btn_container'):
                self.floating_btn_container.place(x=15, y=int(new_y), width=50, height=50)
            
            # Continuar la animación con ritmo más lento para mayor fluidez
            self.after(60, lambda: self._animate_floating(step + 1))
        elif self.floating_menu_visible:
            # Detener animación si el menú está visible
            self.floating_animation_running = False

    def _on_btn_hover_enter(self, event):
        """Efecto al pasar el mouse - vibración y cambio de color"""
        if not self.hover_animation_running:
            self.hover_animation_running = True
            # Cambiar a color más vibrante
            self.floating_btn.configure(bootstyle="warning")
            self._start_hover_vibration()

    def _on_btn_hover_leave(self, event):
        """Efecto al salir el mouse - restaurar estado normal"""
        self.hover_animation_running = False
        # Restaurar color original
        self.floating_btn.configure(bootstyle="info")

    def _start_hover_vibration(self):
        """Animación de vibración cuando el mouse está encima"""
        self._vibrate_button(0)

    def _vibrate_button(self, step):
        """Efecto de vibración muy sutil para botón pequeño"""
        if self.hover_animation_running and step < 12:  # Vibración más corta para botón pequeño
            import random
            # Vibración muy sutil para botón compacto (±1 pixel)
            offset_x = random.randint(-1, 1)
            offset_y = random.randint(-1, 1)
            
            base_x = 15 + offset_x
            base_y = self.floating_btn_base_y + offset_y
            
            if hasattr(self, 'floating_btn_container'):
                self.floating_btn_container.place(x=base_x, y=base_y, width=50, height=50)
            
            # Continuar vibración con ritmo más rápido para efecto de tembleque
            self.after(30, lambda: self._vibrate_button(step + 1))
        elif self.hover_animation_running:
            # Reiniciar vibración para efecto continuo mientras hay hover
            self.after(100, lambda: self._vibrate_button(0))

    def _hide_floating_button(self):
        """Ocultar el botón flotante con animación de desvanecimiento"""
        if hasattr(self, 'floating_btn_container'):
            # Detener animaciones actuales
            self.floating_animation_running = False
            self.hover_animation_running = False
            
            # Animación de desvanecimiento (escala hacia 0)
            def fade_out(step=0):
                if step <= 10:
                    # Reducir tamaño gradualmente
                    scale = 1 - (step / 10)
                    new_width = int(50 * scale)
                    new_height = int(50 * scale)
                    
                    if new_width > 0 and new_height > 0:
                        # Centrar el botón mientras se reduce
                        offset_x = (50 - new_width) // 2
                        offset_y = (50 - new_height) // 2
                        current_y = self.floating_btn_base_y if hasattr(self, 'floating_btn_base_y') else 150
                        
                        self.floating_btn_container.place(
                            x=15 + offset_x, 
                            y=current_y + offset_y, 
                            width=new_width, 
                            height=new_height
                        )
                    
                    self.after(20, lambda: fade_out(step + 1))
                else:
                    # Ocultar completamente
                    self.floating_btn_container.place_forget()
            
            fade_out()

    def _show_floating_button(self):
        """Mostrar el botón flotante con animación de aparición"""
        if hasattr(self, 'floating_btn_container'):
            # Determinar posición correcta según si el header está visible
            current_y = self.floating_btn_base_y if hasattr(self, 'floating_btn_base_y') else 150
            
            # Animación de aparición (escala desde 0)
            def fade_in(step=0):
                if step <= 10:
                    # Aumentar tamaño gradualmente
                    scale = step / 10
                    new_width = int(50 * scale)
                    new_height = int(50 * scale)
                    
                    if new_width > 0 and new_height > 0:
                        # Centrar el botón mientras crece
                        offset_x = (50 - new_width) // 2
                        offset_y = (50 - new_height) // 2
                        
                        self.floating_btn_container.place(
                            x=15 + offset_x, 
                            y=current_y + offset_y, 
                            width=new_width, 
                            height=new_height
                        )
                    
                    self.after(20, lambda: fade_in(step + 1))
                else:
                    # Restaurar tamaño completo y reactivar animaciones
                    self.floating_btn_container.place(x=15, y=current_y, width=50, height=50)
                    self._start_floating_animation()
            
            fade_in()

    def _toggle_floating_menu(self):
        """Alternar visibilidad del menú flotante con animación"""
        if self.floating_menu_visible:
            self._hide_floating_menu()
        else:
            self._show_floating_menu()

    def _show_floating_menu(self):
        """Mostrar el menú flotante con animación suave y ocultar botón"""
        self.floating_menu_visible = True
        
        # Ocultar el botón flotante con animación de desvanecimiento
        self._hide_floating_button()
        
        # Animación suave con easing (ease-out)
        def slide_in(step=0):
            if step <= 25:
                # Función de easing ease-out cuádrica
                progress = step / 25
                eased_progress = 1 - (1 - progress) ** 2
                x_pos = -300 + (eased_progress * 300)  # De -300 a 0
                
                self.floating_menu.place(x=int(x_pos), y=20, width=280, height=420)
                self.after(12, lambda: slide_in(step + 1))
        
        slide_in()

    def _hide_floating_menu(self):
        """Ocultar el menú flotante con animación suave y mostrar botón"""
        self.floating_menu_visible = False
        
        # Animación suave con easing (ease-in)
        def slide_out(step=0):
            if step <= 25:
                # Función de easing ease-in cuádrica
                progress = step / 25
                eased_progress = progress ** 2
                x_pos = 0 - (eased_progress * 300)  # De 0 a -300
                
                self.floating_menu.place(x=int(x_pos), y=20, width=280, height=420)
                self.after(12, lambda: slide_out(step + 1))
            else:
                # Cuando la animación termine, mostrar el botón flotante
                self._show_floating_button()
        
        slide_out()

    # Funciones de navegación modernas
    def _nav_to_welcome(self):
        """Navegar a la pantalla de bienvenida"""
        self._hide_floating_menu()
        self.current_view = "welcome"
        self._show_header()  # Mostrar header solo en bienvenida
        self.show_welcome_screen()

    def _nav_to_database(self):
        """Navegar a la sección de base de datos"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        self.current_view = "database"
        self._show_panel(self.database_frame)

    def _nav_to_visualizar(self):
        """Navegar a la sección de visualización de datos"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        self.current_view = "visualizar"
        self._show_panel(self.visualizar_frame)

    def _nav_to_dashboard(self):
        """Navegar a la sección de dashboard"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        self.current_view = "dashboard"
        self._show_panel(self.dashboard_frame)

    def _nav_to_web_auto(self):
        """Navegar a la sección de automatización web"""
        self._hide_floating_menu()
        self._hide_header_if_not_welcome()
        messagebox.showinfo("Web Automation", "Función de automatización web - En desarrollo")

    def _show_version_info(self):
        """Mostrar información detallada de la versión del sistema"""
        self._hide_floating_menu()
        
        try:
            version_info = get_full_version_info()
            actual_deps = get_dependencies_actual()
            
            # Crear ventana modal
            version_window = ttk.Toplevel(self)
            version_window.title(f"Acerca de - {version_info['project']['name']}")
            version_window.geometry("900x800")
            version_window.resizable(True, True)
            version_window.transient(self)
            version_window.grab_set()
            
            # Configurar estilo de ventana
            version_window.configure(bg='white')
            
            # Centrar la ventana
            version_window.update_idletasks()
            x = (version_window.winfo_screenwidth() // 2) - (900 // 2)
            y = (version_window.winfo_screenheight() // 2) - (800 // 2)
            version_window.geometry(f"900x800+{x}+{y}")
            
            # Frame principal
            main_frame = ttk.Frame(version_window, padding=10)
            main_frame.pack(fill=BOTH, expand=True)
            
            # Header con información principal
            header_frame = ttk.Frame(main_frame, bootstyle="primary", padding=15)
            header_frame.pack(fill=X, pady=(0, 10))
            
            ttk.Label(
                header_frame,
                text=version_info['project']['name'],
                font=("Arial", 18, "bold"),
                bootstyle="inverse-primary"
            ).pack()
            
            ttk.Label(
                header_frame,
                text=f"{get_version_string()} | {get_build_info()}",
                font=("Arial", 12),
                bootstyle="inverse-primary"
            ).pack(pady=(5, 0))
            
            ttk.Label(
                header_frame,
                text=version_info['project']['description'],
                font=("Arial", 10),
                bootstyle="inverse-primary"
            ).pack(pady=(5, 0))
            
            # Notebook para las diferentes secciones
            notebook = ttk.Notebook(main_frame)
            notebook.pack(fill=BOTH, expand=True, pady=10)
            
            # Tab 1: Información General
            info_frame = ttk.Frame(notebook, padding=10)
            notebook.add(info_frame, text="📋 General")
            
            # Crear scroll para info_frame
            info_canvas = tk.Canvas(info_frame)
            info_scrollbar = ttk.Scrollbar(info_frame, orient="vertical", command=info_canvas.yview)
            info_scrollable = ttk.Frame(info_canvas)
            
            info_scrollable.bind(
                "<Configure>",
                lambda e: info_canvas.configure(scrollregion=info_canvas.bbox("all"))
            )
            
            info_canvas.create_window((0, 0), window=info_scrollable, anchor="nw")
            info_canvas.configure(yscrollcommand=info_scrollbar.set)
            
            info_canvas.pack(side="left", fill="both", expand=True)
            info_scrollbar.pack(side="right", fill="y")
            
            self._create_info_section(info_scrollable, "Información del Proyecto", [
                ("Nombre Completo", version_info['project']['full_name']),
                ("Organización", version_info['project']['organization']),
                ("Versión", version_info['version']['version']),
                ("Nombre de Versión", version_info['version']['version_name']),
                ("Tipo de Release", version_info['version']['release_type']),
                ("Código", version_info['version']['codename']),
                ("Fecha de Build", version_info['version']['build_date']),
                ("Número de Build", version_info['version']['build_number']),
                ("Licencia", version_info['project']['license']),
                ("Repositorio", version_info['project']['repository'])
            ])
            
            # Tab 2: Sistema
            system_frame = ttk.Frame(notebook, padding=10)
            notebook.add(system_frame, text="💻 Sistema")
            
            # Crear scroll para system_frame
            system_canvas = tk.Canvas(system_frame)
            system_scrollbar = ttk.Scrollbar(system_frame, orient="vertical", command=system_canvas.yview)
            system_scrollable = ttk.Frame(system_canvas)
            
            system_scrollable.bind(
                "<Configure>",
                lambda e: system_canvas.configure(scrollregion=system_canvas.bbox("all"))
            )
            
            system_canvas.create_window((0, 0), window=system_scrollable, anchor="nw")
            system_canvas.configure(yscrollcommand=system_scrollbar.set)
            
            system_canvas.pack(side="left", fill="both", expand=True)
            system_scrollbar.pack(side="right", fill="y")
            
            # Información básica del sistema
            basic_system_info = [
                ("Versión Python", version_info['system']['python_version'].split()[0]),
                ("Plataforma", version_info['system']['platform']),
                ("Sistema", version_info['system'].get('system', 'No disponible')),
                ("Release", version_info['system'].get('release', 'No disponible')),
                ("Arquitectura", version_info['system']['architecture']),
                ("Máquina", version_info['system'].get('machine', 'No disponible')),
                ("Nodo", version_info['system'].get('node', 'No disponible')),
                ("Procesador", version_info['system']['processor'] or "No disponible")
            ]
            
            self._create_info_section(system_scrollable, "Información Básica", basic_system_info)
            
            # Información de memoria si está disponible
            if 'memoria_total' in version_info['system']:
                memory_info = [
                    ("Memoria Total", version_info['system']['memoria_total']),
                    ("Memoria Disponible", version_info['system']['memoria_disponible']),
                    ("Memoria Usada", version_info['system']['memoria_usada']),
                    ("Porcentaje Usado", version_info['system']['memoria_porcentaje'])
                ]
                self._create_info_section(system_scrollable, "Información de Memoria", memory_info)
            
            # Información de CPU si está disponible
            if 'cpu_cores' in version_info['system']:
                cpu_info = [
                    ("Núcleos Físicos", str(version_info['system']['cpu_cores'])),
                    ("Hilos Lógicos", str(version_info['system']['cpu_threads'])),
                    ("Frecuencia Máxima", version_info['system']['cpu_frecuencia'])
                ]
                self._create_info_section(system_scrollable, "Información del Procesador", cpu_info)
            
            # Información de hardware adicional
            hardware_info = []
            if 'tarjeta_grafica' in version_info['system']:
                gpus = version_info['system']['tarjeta_grafica']
                if isinstance(gpus, list):
                    for i, gpu in enumerate(gpus):
                        hardware_info.append((f"Tarjeta Gráfica {i+1}", gpu))
                else:
                    hardware_info.append(("Tarjeta Gráfica", str(gpus)))
            
            if 'placa_madre' in version_info['system']:
                hardware_info.append(("Placa Madre", version_info['system']['placa_madre']))
            
            if hardware_info:
                self._create_info_section(system_scrollable, "Hardware", hardware_info)
            
            # Información de discos si está disponible
            if 'discos' in version_info['system'] and isinstance(version_info['system']['discos'], list):
                for i, disco in enumerate(version_info['system']['discos']):
                    if isinstance(disco, dict):
                        disk_info = [
                            ("Dispositivo", disco.get('dispositivo', 'No disponible')),
                            ("Punto de Montaje", disco.get('punto_montaje', 'No disponible')),
                            ("Sistema de Archivos", disco.get('sistema_archivos', 'No disponible')),
                            ("Espacio Total", disco.get('total', 'No disponible')),
                            ("Espacio Usado", disco.get('usado', 'No disponible')),
                            ("Espacio Libre", disco.get('libre', 'No disponible')),
                            ("Porcentaje Usado", disco.get('porcentaje', 'No disponible'))
                        ]
                        self._create_info_section(system_scrollable, f"Disco {i+1}", disk_info)
            
            # Tab 3: Dependencias
            deps_frame = ttk.Frame(notebook, padding=10)
            notebook.add(deps_frame, text="📦 Dependencias")
            
            # Crear tabla de dependencias
            deps_tree_frame = ttk.Frame(deps_frame)
            deps_tree_frame.pack(fill=BOTH, expand=True)
            
            deps_tree = ttk.Treeview(
                deps_tree_frame,
                columns=("esperada", "actual", "estado"),
                show="tree headings",
                selectmode="extended"
            )
            
            # Configurar columnas
            deps_tree.heading("#0", text="Paquete")
            deps_tree.heading("esperada", text="Versión Esperada")
            deps_tree.heading("actual", text="Versión Actual")
            deps_tree.heading("estado", text="Estado")
            
            deps_tree.column("#0", width=150, minwidth=100)
            deps_tree.column("esperada", width=150, minwidth=100)
            deps_tree.column("actual", width=150, minwidth=100)
            deps_tree.column("estado", width=100, minwidth=80)
            
            # Llenar dependencias
            for package, expected_version in version_info['dependencies'].items():
                actual_version = actual_deps.get(package, "❌ No instalado")
                
                if actual_version.startswith("❌"):
                    status = "❌ No instalado"
                elif actual_version.startswith("⚠️"):
                    status = "⚠️ Error"
                elif actual_version.startswith("✅"):
                    status = "✅ Instalado"
                else:
                    status = "✅ OK"
                
                deps_tree.insert(
                    "",
                    "end",
                    text=package,
                    values=(expected_version, actual_version, status)
                )
            
            deps_tree.pack(fill=BOTH, expand=True)
            
            # Scrollbar para el treeview
            deps_scrollbar = ttk.Scrollbar(deps_tree_frame, orient="vertical", command=deps_tree.yview)
            deps_tree.configure(yscrollcommand=deps_scrollbar.set)
            deps_scrollbar.pack(side="right", fill="y")
            
            # Tab 4: Equipo de Desarrollo
            team_frame = ttk.Frame(notebook, padding=10)
            notebook.add(team_frame, text="👥 Equipo")
            
            # Crear scroll para team_frame
            team_canvas = tk.Canvas(team_frame)
            team_scrollbar = ttk.Scrollbar(team_frame, orient="vertical", command=team_canvas.yview)
            team_scrollable = ttk.Frame(team_canvas)
            
            team_scrollable.bind(
                "<Configure>",
                lambda e: team_canvas.configure(scrollregion=team_canvas.bbox("all"))
            )
            
            team_canvas.create_window((0, 0), window=team_scrollable, anchor="nw")
            team_canvas.configure(yscrollcommand=team_scrollbar.set)
            
            team_canvas.pack(side="left", fill="both", expand=True)
            team_scrollbar.pack(side="right", fill="y")
            
            # Información del equipo
            role_titles = {
                'desarrollador': '👨‍💻 Desarrollador',
                'lider_investigacion': '👨‍⚕️ Líder de Investigación y Proyección Oncológica',
                'jefe_gestion_informacion': '👨‍💼 Jefe de Gestión de la Información'
            }
            
            for role_key, role_info in version_info['team'].items():
                role_data = [
                    ("Nombre", role_info['nombre']),
                    ("Cargo", role_info['cargo']),
                    ("Departamento", role_info['departamento']),
                    ("Correo", role_info['correo'])
                ]
                title = role_titles.get(role_key, role_info['cargo'])
                self._create_info_section(team_scrollable, title, role_data)
            
            # Tab 5: Características
            features_frame = ttk.Frame(notebook, padding=10)
            notebook.add(features_frame, text="✨ Características")
            
            features_text = ttk.Text(features_frame, wrap=WORD, height=15, font=("Consolas", 10))
            features_text.pack(fill=BOTH, expand=True)
            
            features_content = "🔥 CARACTERÍSTICAS DE LA VERSIÓN ACTUAL:\n\n"
            for feature in version_info['features']:
                features_content += f"{feature}\n"
            
            features_content += "\n\n📊 MÉTRICAS DE RENDIMIENTO:\n\n"
            for metric, value in version_info['performance'].items():
                features_content += f"• {metric.replace('_', ' ').title()}: {value}\n"
            
            features_content += "\n\n👥 AUDIENCIAS OBJETIVO:\n\n"
            for audience, benefit in version_info['audiences'].items():
                features_content += f"• {audience}: {benefit}\n"
            
            features_text.insert("1.0", features_content)
            features_text.configure(state="disabled")
            
            # Tab 6: Roadmap
            roadmap_frame = ttk.Frame(notebook, padding=10)
            notebook.add(roadmap_frame, text="🗺️ Roadmap")
            
            roadmap_text = ttk.Text(roadmap_frame, wrap=WORD, height=15, font=("Consolas", 10))
            roadmap_text.pack(fill=BOTH, expand=True)
            
            roadmap_content = "🚀 PRÓXIMAS VERSIONES:\n\n"
            for version, description in version_info['roadmap'].items():
                roadmap_content += f"{version}: {description}\n\n"
            
            roadmap_text.insert("1.0", roadmap_content)
            roadmap_text.configure(state="disabled")
            
            # Frame de botones
            buttons_frame = ttk.Frame(main_frame, padding=10)
            buttons_frame.pack(fill=X, pady=10)
            
            ttk.Button(
                buttons_frame,
                text="📋 Copiar Info Sistema",
                command=lambda: self._copy_system_info_to_clipboard(version_info),
                bootstyle="info"
            ).pack(side=LEFT, padx=(0, 10))
            
            ttk.Button(
                buttons_frame,
                text="✅ Cerrar",
                command=version_window.destroy,
                bootstyle="success"
            ).pack(side=RIGHT)
            
            # Bind mouse wheel para scroll en cada canvas
            def _on_mousewheel(event):
                try:
                    # Determinar qué canvas está activo basado en el widget con focus
                    widget = event.widget
                    while widget:
                        if isinstance(widget, tk.Canvas):
                            widget.yview_scroll(int(-1*(event.delta/120)), "units")
                            break
                        widget = widget.master
                except:
                    pass
            
            # Bind a la ventana completa
            version_window.bind_all("<MouseWheel>", _on_mousewheel)
            
            # Cleanup al cerrar
            def on_closing():
                try:
                    version_window.unbind_all("<MouseWheel>")
                except:
                    pass
                version_window.destroy()
            
            version_window.protocol("WM_DELETE_WINDOW", on_closing)
            
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"Error al mostrar información de versión:\n{str(e)}"
            )

    def _create_info_section(self, parent, title, info_items):
        """Crear una sección de información con título y elementos"""
        # Frame para la sección
        section_frame = ttk.LabelFrame(parent, text=title, padding=10)
        section_frame.pack(fill=X, pady=(0, 10))
        
        # Grid de información
        for i, (label, value) in enumerate(info_items):
            ttk.Label(
                section_frame,
                text=f"{label}:",
                font=("Arial", 9, "bold")
            ).grid(row=i, column=0, sticky=W, padx=(0, 10), pady=2)
            
            ttk.Label(
                section_frame,
                text=str(value),
                font=("Arial", 9)
            ).grid(row=i, column=1, sticky=W, pady=2)
    
    def _copy_system_info_to_clipboard(self, version_info):
        """Copiar información del sistema al clipboard"""
        try:
            info_text = f"""EVARISIS Gestor H.U.V - Información del Sistema
=====================================
Versión: {get_version_string()}
Build: {get_build_info()}
Python: {version_info['system']['python_version'].split()[0]}
Plataforma: {version_info['system']['platform']}
Arquitectura: {version_info['system']['architecture']}

Dependencias:
"""
            actual_deps = get_dependencies_actual()
            for package, expected in version_info['dependencies'].items():
                actual = actual_deps.get(package, "No instalado")
                info_text += f"- {package}: {actual} (esperado: {expected})\n"
            
            self.clipboard_clear()
            self.clipboard_append(info_text)
            self.update()  # Actualizar para que el clipboard se procese
            
            messagebox.showinfo("Copiado", "Información del sistema copiada al portapapeles")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al copiar al portapapeles:\n{str(e)}")

    def _hide_header_if_not_welcome(self):
        """Ocultar header si no estamos en la pantalla de bienvenida"""
        if self.header_visible:
            self.header.pack_forget()
            self.header_visible = False
            # Actualizar posición base del botón flotante sin header
            self.floating_btn_base_y = 20
            # Reposicionar inmediatamente el botón si existe
            if hasattr(self, 'floating_btn_container') and not self.floating_menu_visible:
                self.floating_btn_container.place(x=15, y=20, width=50, height=50)

    def _show_header(self):
        """Mostrar header (solo para pantalla de bienvenida)"""
        if not self.header_visible:
            self.header.pack(fill=X, before=self.content_container)
            self.header_visible = True
            # Reajustar posición base del botón flotante con header visible
            self.floating_btn_base_y = 150
            # Reposicionar inmediatamente el botón
            if hasattr(self, 'floating_btn_container'):
                self.floating_btn_container.place(x=15, y=150, width=50, height=50)
            # Reposicionar el botón flotante cuando se muestra el header
            self.floating_btn.place(x=20, y=20)

    def _create_sidebar(self):
        """Crear la barra lateral de navegación"""
        # Header del sidebar elegante
        top = ttk.Frame(self.sidebar, padding=10)
        top.pack(fill=X)
        
        # Título del sidebar
        ttk.Label(
            top, 
            text="� NAVEGACIÓN", 
            font=("Segoe UI", 12, "bold"),
            anchor="center"
        ).pack(fill=X, pady=(0, 10))

        # Navegación
        nav = ttk.Frame(self.sidebar, padding=(10, 10))
        nav.pack(fill=BOTH, expand=True)

        self.nav_buttons = {}

        # Botones de navegación reorganizados
        nav_items = [
            ("🏠 Inicio", "home", "light", self.show_welcome_screen),
            ("🗄️ Base de Datos", "database", "primary", self.show_database_frame),
            ("📊 Visualizar Datos", "informes", "success", self.show_visualizar_frame),
            ("📈 Análisis Gráfico", "dashboard", "info", self.show_dashboard_frame),
            ("🔄 Automatizador Fuente PDF\n(Experimental)", "web", "warning", self.open_web_auto_modal),
        ]

        for text, icon_key, style, callback in nav_items:
            btn = ttk.Button(
                nav, 
                text=text,
                image=self.iconos.get(icon_key),
                compound=LEFT,
                bootstyle=style,
                command=callback,
                width=20
            )
            btn.pack(fill=X, pady=2)
            self.nav_buttons[text] = btn

        # Botón de navegación (mostrar/ocultar menús)
        self.nav_toggle_btn = ttk.Button(
            nav, 
            text="◀ Ocultar Menús", 
            command=self._toggle_navigation_visibility, 
            bootstyle="secondary",
            width=20
        )
        self.nav_toggle_btn.pack(fill=X, pady=(20, 0))

        # Footer
        ttk.Label(
            self.sidebar, 
            text="HUV • EVARISIS", 
            anchor=CENTER, 
            padding=(10, 8), 
            bootstyle="light"
        ).pack(side=BOTTOM, fill=X)

    def _toggle_sidebar(self):
        """Alternar visibilidad de la sidebar"""
        target = 0 if self.sidebar_expanded else self.sidebar_width
        step = -24 if self.sidebar_expanded else 24

        def animate(curr):
            if (step > 0 and curr < target) or (step < 0 and curr > target):
                curr += step
                self.sidebar.configure(width=max(0, curr))
                self.after(10, lambda: animate(curr))
            else:
                self.sidebar.configure(width=target)
                self.sidebar_expanded = not self.sidebar_expanded

        current = self.sidebar.winfo_width() or (self.sidebar_width if self.sidebar_expanded else 0)
        animate(current)

    def _create_content_panels(self):
        """Crear los paneles de contenido principal con scroll"""
        # Panel de base de datos con scroll
        self.database_frame = self._create_scrollable_frame(self.content_container)
        self._create_database_content()

        # Panel de visualización con scroll
        self.visualizar_frame = self._create_scrollable_frame(self.content_container)
        self._create_visualizar_content()

        # Panel de análisis gráfico con scroll
        self.dashboard_frame = self._create_scrollable_frame(self.content_container)
        self._create_dashboard_content()

        # Panel activo actual
        self.panel_activo = None

    def _create_scrollable_frame(self, parent):
        """Crear un frame con barra de desplazamiento"""
        # Frame contenedor principal que llena toda el área
        container = ttk.Frame(parent, padding=0)
        
        # Canvas para scroll sin bordes ni highlight
        canvas = tk.Canvas(container, highlightthickness=0, borderwidth=0, relief='flat')
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, padding=10)
        
        # Configurar scroll
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Hacer que el scrollable_frame se expanda horizontalmente en el canvas
        def _configure_canvas(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Hacer que el frame interno sea del mismo ancho que el canvas
            canvas_width = event.width
            canvas.itemconfig(window_id, width=canvas_width)
        
        window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.bind('<Configure>', _configure_canvas)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Scroll con rueda del mouse - vinculado específicamente al canvas y sus hijos
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        # Vincular el evento al canvas específico
        canvas.bind("<MouseWheel>", _on_mousewheel)
        
        # También vincular a todos los widgets hijos del scrollable_frame
        def _bind_to_mousewheel(widget):
            # No vincular scroll a widgets que ya manejan su propio scroll
            widget_type = widget.winfo_class()
            if widget_type not in ['Listbox', 'Treeview', 'Text']:
                widget.bind("<MouseWheel>", _on_mousewheel)
            
            for child in widget.winfo_children():
                _bind_to_mousewheel(child)
        
        # Empacar elementos para llenar toda el área
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Aplicar el binding del scroll a todos los widgets después de un pequeño delay
        # para asegurar que todos los widgets hijos se han creado
        def _apply_scroll_binding():
            try:
                _bind_to_mousewheel(scrollable_frame)
            except:
                pass  # Si hay error, continuamos sin problemas
        
        container.after(100, _apply_scroll_binding)
        
        # Guardar referencias para uso posterior
        container.scrollable_frame = scrollable_frame
        container.canvas = canvas
        container._bind_to_mousewheel = _bind_to_mousewheel  # Guardar función para uso posterior
        
        return container

    def _create_database_content(self):
        """Crear contenido del panel de base de datos"""
        # Usar el frame scrollable
        frame = self.database_frame.scrollable_frame
        
        # Título principal
        ttk.Label(
            frame, 
            text="🗄️ Estado de la Base de Datos", 
            font=self.FONT_TITULO
        ).pack(pady=(0, 20), anchor=W)

        # Dashboard de estadísticas
        stats_container = ttk.Frame(frame)
        stats_container.pack(expand=True, fill=BOTH)
        stats_container.grid_columnconfigure(0, weight=1)
        stats_container.grid_columnconfigure(1, weight=1)
        stats_container.grid_rowconfigure(0, weight=1)
        
        # Panel izquierdo - Estadísticas generales
        left_panel = ttk.Frame(stats_container, padding=20)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        
        # Crear cards de estadísticas
        self._create_stats_cards(left_panel)
        
        # Panel derecho - Importación y procesamiento
        right_panel = ttk.Frame(stats_container, padding=20)
        right_panel.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        
        # Sección de importación
        import_section = ttk.LabelFrame(right_panel, text="📥 Importar Nuevos Datos", padding=15)
        import_section.pack(fill=X, pady=(0, 20))
        
        ttk.Button(
            import_section, 
            text="📄 Agregar Archivo PDF", 
            command=self._select_pdf_file,
            bootstyle="primary",
            width=25
        ).pack(fill=X, pady=5)
        
        ttk.Button(
            import_section, 
            text="📁 Importar Carpeta de PDFs", 
            command=self._select_pdf_folder,
            bootstyle="primary",
            width=25
        ).pack(fill=X, pady=5)
        
        # Sección de archivos disponibles
        files_section = ttk.LabelFrame(right_panel, text="📂 Archivos Disponibles", padding=15)
        files_section.pack(expand=True, fill=BOTH)
        
        # Lista de archivos con scrollbar
        list_frame = ttk.Frame(files_section)
        list_frame.pack(expand=True, fill=BOTH, pady=(0, 10))
        
        self.files_listbox = tk.Listbox(
            list_frame,
            selectmode=tk.EXTENDED,
            font=("Segoe UI", 9),
            height=8
        )
        self.files_listbox.pack(side=LEFT, expand=True, fill=BOTH)
        
        files_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.files_listbox.yview)
        files_scrollbar.pack(side=RIGHT, fill=Y)
        self.files_listbox.configure(yscrollcommand=files_scrollbar.set)
        
        # Botones de control
        control_frame = ttk.Frame(files_section)
        control_frame.pack(fill=X)
        control_frame.grid_columnconfigure(0, weight=1)
        control_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Button(
            control_frame, 
            text="🔄 Actualizar", 
            command=self._refresh_files_list,
            bootstyle="secondary"
        ).grid(row=0, column=0, sticky="ew", padx=(0, 5))
        
        ttk.Button(
            control_frame, 
            text="⚡ Procesar Seleccionados", 
            command=self._process_selected_files,
            bootstyle="success"
        ).grid(row=0, column=1, sticky="ew", padx=(5, 0))

        # Inicializar datos
        self._refresh_files_list()
        self._refresh_database_stats()

    def _create_visualizar_content(self):
        """Crear contenido del panel de visualización mejorado"""
        # Usar el frame scrollable
        frame = self.visualizar_frame.scrollable_frame
        
        # Título principal
        title_frame = ttk.Frame(frame)
        title_frame.pack(fill=X, pady=(0, 15))
        
        ttk.Label(
            title_frame, 
            text="📊 Visualizador de Datos", 
            font=self.FONT_TITULO
        ).pack(side=LEFT)
        
        # Botones de acción en el header
        actions_frame = ttk.Frame(title_frame)
        actions_frame.pack(side=RIGHT)
        
        ttk.Button(
            actions_frame, 
            text="📤 Exportar Selección", 
            command=self._export_selected_data,
            bootstyle="success"
        ).pack(side=RIGHT, padx=(10, 0))
        
        ttk.Button(
            actions_frame, 
            text="🔄 Actualizar Datos", 
            command=self.refresh_data_and_table,
            bootstyle="primary"
        ).pack(side=RIGHT)
        
        # Container principal con layout en grid
        main_container = ttk.Frame(frame)
        main_container.pack(expand=True, fill=BOTH)
        main_container.grid_columnconfigure(0, weight=1)  # tabla selector (más pequeña)
        main_container.grid_columnconfigure(1, weight=2)  # tabla detalles (más grande)
        main_container.grid_rowconfigure(1, weight=1)

        # Frame para tabla (lado izquierdo)
        table_frame = ttk.Frame(main_container)
        table_frame.grid(row=0, column=0, rowspan=2, sticky="nsew", padx=(0, 10))
        table_frame.grid_rowconfigure(1, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Campo de búsqueda
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.filter_tabla)
        search_entry = ttk.Entry(
            table_frame,
            textvariable=self.search_var,
            font=("Segoe UI", 11)
        )
        search_entry.grid(row=0, column=0, sticky="ew", padx=10, pady=10)
        search_entry.insert(0, "Buscar por N° Petición, Nombre o Apellido...")

        # Configurar el estilo del Treeview
        style = self.setup_treeview_style()
        
        # Treeview para mostrar datos
        self.tree = ttk.Treeview(table_frame, show="headings", style="Custom.Treeview")
        self.tree.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))
        self.tree.bind("<<TreeviewSelect>>", self.mostrar_detalle_registro)

        # Scrollbar para el Treeview
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=1, column=1, sticky="ns", pady=(0, 10))
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Frame para detalles (lado derecho)
        self.detail_frame = ttk.Frame(main_container, padding=10)
        self.detail_frame.grid(row=0, column=1, rowspan=2, sticky="nsew")
        self.detail_frame.grid_columnconfigure(0, weight=1)
        self.detail_frame.grid_rowconfigure(1, weight=1)

        # Título del panel de detalles
        ttk.Label(
            self.detail_frame, 
            text="Detalles del Registro", 
            font=("Segoe UI", 16, "bold")
        ).grid(row=0, column=0, pady=(0, 10), sticky="w")
        
        # Área de texto para detalles
        self.detail_textbox = tk.Text(
            self.detail_frame, 
            state="disabled", 
            wrap="word", 
            font=("Calibri", 14),
            relief="flat",
            padx=10,
            pady=10
        )
        self.detail_textbox.grid(row=1, column=0, sticky="nsew")
        
        # Scrollbar para el área de texto
        detail_scrollbar = ttk.Scrollbar(self.detail_frame, orient="vertical", command=self.detail_textbox.yview)
        detail_scrollbar.grid(row=1, column=1, sticky="ns")
        self.detail_textbox.configure(yscrollcommand=detail_scrollbar.set)

    def _create_dashboard_content(self):
        """Crear contenido del panel de dashboard"""
        # Usar el frame scrollable
        frame = self.dashboard_frame.scrollable_frame
        
        # Título principal actualizado
        ttk.Label(frame, text="📈 Análisis Gráfico de la Base de Datos", font=self.FONT_TITULO).pack(pady=(0, 10), anchor=W)
        
        # Container principal
        main_container = ttk.Frame(frame)
        main_container.pack(expand=True, fill=BOTH)
        main_container.grid_rowconfigure(0, weight=1)
        main_container.grid_columnconfigure(0, weight=0)  # sidebar
        main_container.grid_columnconfigure(1, weight=1)  # main area

        # Sidebar de filtros (inicialmente oculto)
        self.db_filters = {
            "fecha_desde": tk.StringVar(value=""),
            "fecha_hasta": tk.StringVar(value=""),
            "servicio": tk.StringVar(value=""),
            "malignidad": tk.StringVar(value=""),
            "responsable": tk.StringVar(value=""),
        }
        self.db_sidebar_collapsed = True
        self.db_sidebar = ttk.Frame(main_container, padding=15, width=280)
        
        # Título del sidebar de filtros
        ttk.Label(self.db_sidebar, text="Filtros", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 10))
        
        # Campos de filtros
        ttk.Label(self.db_sidebar, text="Fecha desde (dd/mm/aaaa)").grid(row=1, column=0, sticky="w", pady=(5, 2))
        ttk.Entry(self.db_sidebar, textvariable=self.db_filters["fecha_desde"]).grid(row=2, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Fecha hasta (dd/mm/aaaa)").grid(row=3, column=0, sticky="w", pady=(5, 2))
        ttk.Entry(self.db_sidebar, textvariable=self.db_filters["fecha_hasta"]).grid(row=4, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Servicio").grid(row=5, column=0, sticky="w", pady=(5, 2))
        self.cmb_servicio = ttk.Combobox(self.db_sidebar, textvariable=self.db_filters["servicio"], values=[])
        self.cmb_servicio.grid(row=6, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Malignidad").grid(row=7, column=0, sticky="w", pady=(5, 2))
        self.cmb_malig = ttk.Combobox(self.db_sidebar, textvariable=self.db_filters["malignidad"], values=["", "PRESENTE", "AUSENTE"])
        self.cmb_malig.grid(row=8, column=0, sticky="ew", pady=(0, 10))
        
        ttk.Label(self.db_sidebar, text="Responsable").grid(row=9, column=0, sticky="w", pady=(5, 2))
        self.cmb_resp = ttk.Combobox(self.db_sidebar, textvariable=self.db_filters["responsable"], values=[])
        self.cmb_resp.grid(row=10, column=0, sticky="ew", pady=(0, 10))
        
        # Botones del sidebar
        btns_frame = ttk.Frame(self.db_sidebar)
        btns_frame.grid(row=11, column=0, sticky="ew", pady=(10, 0))
        btns_frame.grid_columnconfigure(0, weight=1)
        btns_frame.grid_columnconfigure(1, weight=1)
        
        ttk.Button(btns_frame, text="Refrescar", command=self._refresh_dashboard).grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(btns_frame, text="Limpiar", command=self._clear_filters).grid(row=0, column=1, sticky="ew", padx=(5, 0))
        
        # Configurar grid del sidebar
        self.db_sidebar.grid_columnconfigure(0, weight=1)

        # Área principal con toolbar + notebook
        main_area = ttk.Frame(main_container)
        main_area.grid(row=0, column=1, sticky="nsew")
        main_area.grid_rowconfigure(1, weight=1)
        main_area.grid_columnconfigure(0, weight=1)

        # Toolbar superior
        toolbar = ttk.Frame(main_area, padding=(5, 5))
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        
        self.btn_toggle_sidebar = ttk.Button(toolbar, text="≡ Mostrar filtros", command=self._toggle_db_sidebar)
        self.btn_toggle_sidebar.pack(side=LEFT, padx=(0, 10))
        
        ttk.Button(toolbar, text="Filtros…", command=self._open_filters_sheet).pack(side=LEFT)

        # Notebook con las pestañas del dashboard
        self.tabs = ttk.Notebook(main_area)
        self.tabs.grid(row=1, column=0, sticky="nsew")

        # Crear las pestañas con scroll
        self.tab_overview   = ttk.Frame(self.tabs)
        self.tab_biomarkers = ttk.Frame(self.tabs)
        self.tab_times      = ttk.Frame(self.tabs)
        self.tab_quality    = ttk.Frame(self.tabs)
        self.tab_compare    = ttk.Frame(self.tabs)

        # Configurar cada pestaña para ser responsive con grid 2x2
        for tab in [self.tab_overview, self.tab_biomarkers, self.tab_times, self.tab_quality, self.tab_compare]:
            tab.grid_columnconfigure(0, weight=1)
            tab.grid_columnconfigure(1, weight=1)
            tab.grid_rowconfigure(0, weight=1)
            tab.grid_rowconfigure(1, weight=1)

        self.tabs.add(self.tab_overview,   text="Overview")
        self.tabs.add(self.tab_biomarkers, text="Biomarcadores")
        self.tabs.add(self.tab_times,      text="Tiempos")
        self.tabs.add(self.tab_quality,    text="Calidad")
        self.tabs.add(self.tab_compare,    text="Comparador")

        self._dash_canvases = []

    def _create_kpi_card(self, parent, title, value):
        """Crear una tarjeta KPI usando TTKBootstrap"""
        card = ttk.Frame(parent, padding=10, relief="solid", borderwidth=1)
        
        ttk.Label(card, text=title, font=("Segoe UI", 12)).pack(anchor=W, padx=4, pady=(2, 0))
        value_lbl = ttk.Label(card, text=value, font=("Segoe UI", 26, "bold")).pack(anchor=W, padx=4, pady=(0, 2))
        
        # Store reference for updating
        card.value_lbl = list(card.winfo_children())[-1] if card.winfo_children() else None
        
        return card

    def _create_welcome_screen(self):
        """Crear la pantalla de bienvenida inicial"""
        self.welcome_frame = ttk.Frame(self.content_container, padding=40)
        
        # Contenedor central
        center_container = ttk.Frame(self.welcome_frame)
        center_container.pack(expand=True, fill=BOTH)
        center_container.grid_rowconfigure(0, weight=1)
        center_container.grid_rowconfigure(1, weight=0)
        center_container.grid_rowconfigure(2, weight=1)
        center_container.grid_columnconfigure(0, weight=1)
        
        # Espaciador superior
        ttk.Frame(center_container).grid(row=0, column=0)
        
        # Contenido principal
        main_content = ttk.Frame(center_container)
        main_content.grid(row=1, column=0, pady=50)
        
        # Título principal con emojis
        title_frame = ttk.Frame(main_content)
        title_frame.pack(pady=(0, 30))
        
        ttk.Label(
            title_frame,
            text="🏥 Bienvenido al Gestor Oncológico 🔬",
            font=("Segoe UI", 28, "bold"),
            anchor="center"
        ).pack()
        
        # Subtítulo descriptivo
        subtitle_text = ("Enfocado en la investigación y mejora del área de oncología\n"
                        "del Hospital Universitario del Valle")
        ttk.Label(
            main_content,
            text=subtitle_text,
            font=("Segoe UI", 16),
            anchor="center",
            justify="center"
        ).pack(pady=(0, 40))
        
        # Iconos representativos
        icons_frame = ttk.Frame(main_content)
        icons_frame.pack(pady=(0, 40))
        
        # Crear una fila de iconos descriptivos
        icons_text = "📊  📈  🧬  💡  📋  🔍  ⚕️"
        ttk.Label(
            icons_frame,
            text=icons_text,
            font=("Segoe UI", 24),
            anchor="center"
        ).pack()
        
        # Mensaje de instrucción
        instruction_text = ("Selecciona una opción del menú lateral para comenzar\n"
                           "a trabajar con los datos oncológicos")
        ttk.Label(
            main_content,
            text=instruction_text,
            font=("Segoe UI", 14),
            anchor="center",
            justify="center",
            foreground="gray"
        ).pack()
        
        # Espaciador inferior
        ttk.Frame(center_container).grid(row=2, column=0)

    def _create_stats_cards(self, parent):
        """Crear las tarjetas de estadísticas de la base de datos"""
        # Grid para las cards
        parent.grid_columnconfigure(0, weight=1)
        
        # Cards de estadísticas principales
        stats_data = [
            ("📊", "Total Registros", "0", "Número total de informes en la base de datos"),
            ("📅", "Rango Temporal", "—", "Fechas de los informes más antiguos y recientes"),
            ("🕒", "Última Importación", "—", "Fecha de la última importación de datos"),
            ("🏥", "Servicios Únicos", "0", "Cantidad de servicios médicos diferentes"),
            ("⚠️", "Casos Malignos", "0", "Informes con diagnóstico de malignidad"),
            ("🔬", "Biomarcadores", "0", "Informes con análisis de biomarcadores")
        ]
        
        for i, (icon, title, value, description) in enumerate(stats_data):
            card = ttk.Frame(parent, padding=15, relief="solid", borderwidth=1)
            card.grid(row=i//2, column=i%2, padx=10, pady=10, sticky="ew")
            
            # Header con icono y título
            header = ttk.Frame(card)
            header.pack(fill=X, pady=(0, 5))
            
            ttk.Label(header, text=icon, font=("Segoe UI", 20)).pack(side=LEFT, padx=(0, 10))
            ttk.Label(header, text=title, font=("Segoe UI", 12, "bold")).pack(side=LEFT)
            
            # Valor principal
            value_label = ttk.Label(card, text=value, font=("Segoe UI", 24, "bold"))
            value_label.pack(anchor=W, pady=(0, 5))
            
            # Descripción
            ttk.Label(
                card, 
                text=description, 
                font=("Segoe UI", 9), 
                foreground="gray",
                wraplength=200
            ).pack(anchor=W)
            
            # Guardar referencia para actualización
            setattr(self, f"stats_card_{i}", value_label)
        
        # Botón para actualizar estadísticas
        update_btn = ttk.Button(
            parent,
            text="🔄 Actualizar Estadísticas",
            command=self._refresh_database_stats,
            bootstyle="info",
            width=25
        )
        update_btn.grid(row=3, column=0, columnspan=2, pady=(20, 0), sticky="ew")

    def _refresh_database_stats(self):
        """Actualizar las estadísticas de la base de datos"""
        try:
            # Consultar datos reales de la base de datos
            from database_manager import get_all_records_as_dataframe
            
            # Obtener todos los registros
            df = get_all_records_as_dataframe()
            
            if df is None or df.empty:
                stats = {
                    'total_records': 0,
                    'date_range': '—',
                    'last_import': '—',
                    'unique_services': 0,
                    'malignant_count': 0,
                    'biomarker_count': 0
                }
            else:
                # Calcular estadísticas reales
                total_records = len(df)
                
                # Buscar columnas de fecha
                fecha_cols = [col for col in df.columns if 'fecha' in col.lower()]
                fecha_min, fecha_max = None, None
                
                for col in fecha_cols:
                    if col in df.columns:
                        try:
                            # Convertir a datetime con formato día/mes/año
                            fechas = pd.to_datetime(df[col], errors='coerce', format='%d/%m/%Y')
                            fechas_validas = fechas.dropna()
                            
                            if not fechas_validas.empty:
                                col_min = fechas_validas.min()
                                col_max = fechas_validas.max()
                                
                                if fecha_min is None or col_min < fecha_min:
                                    fecha_min = col_min
                                if fecha_max is None or col_max > fecha_max:
                                    fecha_max = col_max
                        except:
                            continue
                
                # Formatear rango de fechas
                if fecha_min and fecha_max:
                    date_range = f"{fecha_min.strftime('%d/%m/%Y')} - {fecha_max.strftime('%d/%m/%Y')}"
                    last_import = fecha_max.strftime('%d/%m/%Y')
                else:
                    date_range = '—'
                    last_import = '—'
                
                # Servicios únicos
                servicio_cols = [col for col in df.columns if 'servicio' in col.lower()]
                unique_services = 0
                if servicio_cols:
                    unique_services = df[servicio_cols[0]].nunique() if not df[servicio_cols[0]].isna().all() else 0
                
                # Casos malignos
                malignant_count = 0
                malignidad_cols = [col for col in df.columns if 'malign' in col.lower()]
                if malignidad_cols:
                    malignant_count = (df[malignidad_cols[0]].str.contains('PRESENTE', case=False, na=False)).sum()
                
                # Biomarcadores (buscar columnas con biomarcadores)
                biomarker_cols = [col for col in df.columns if any(bio in col.lower() for bio in ['ki67', 'her2', 're ', 'rp ', 'pdl1'])]
                biomarker_count = 0
                for col in biomarker_cols:
                    biomarker_count += df[col].notna().sum()
                # Evitar doble conteo - usar registros únicos con al menos un biomarcador
                if biomarker_cols:
                    biomarker_count = df[biomarker_cols].notna().any(axis=1).sum()
                
                stats = {
                    'total_records': total_records,
                    'date_range': date_range,
                    'last_import': last_import,
                    'unique_services': unique_services,
                    'malignant_count': malignant_count,
                    'biomarker_count': biomarker_count
                }
            
            # Actualizar las cards si existen
            values = [
                str(stats['total_records']),
                stats['date_range'],
                stats['last_import'], 
                str(stats['unique_services']),
                str(stats['malignant_count']),
                str(stats['biomarker_count'])
            ]
            
            for i, value in enumerate(values):
                card_attr = f"stats_card_{i}"
                if hasattr(self, card_attr):
                    getattr(self, card_attr).configure(text=value)
                    
        except Exception as e:
            print(f"Error actualizando estadísticas: {e}")

    # Métodos de navegación actualizados
    def show_database_frame(self):
        """Mostrar panel de base de datos (función compatibilidad)"""
        self._nav_to_database()
        # Actualizar estadísticas cuando se muestre el panel
        self._refresh_database_stats()

    def show_visualizar_frame(self):
        """Mostrar panel de visualización (función compatibilidad)"""
        self._nav_to_visualizar()

    def show_dashboard_frame(self):
        """Mostrar panel de análisis gráfico (función compatibilidad)"""
        self._nav_to_dashboard()
        # Cargar el dashboard cuando se muestre
        self.cargar_dashboard()
        
    def show_welcome_screen(self):
        """Mostrar pantalla de bienvenida con header visible y menú flotante oculto"""
        # Ocultar otros paneles
        if self.panel_activo:
            self.panel_activo.pack_forget()
        
        # Ocultar menú flotante si está visible
        if self.floating_menu_visible:
            self._hide_floating_menu()
        
        # Mostrar header (solo en pantalla de bienvenida)
        self._show_header()
        
        # Mostrar la pantalla de bienvenida
        self.welcome_frame.pack(fill=BOTH, expand=True)
        self.panel_activo = self.welcome_frame
        self.welcome_screen_active = True
        self.current_view = "welcome"

    def _hide_welcome_and_show_panel(self, panel):
        """Ocultar pantalla de bienvenida, animar menús y mostrar panel"""
        # Si es la primera navegación desde la pantalla de bienvenida
        if self.welcome_screen_active:
            self._animate_menus_hide()
            self.welcome_screen_active = False
        
        # Cambiar el panel
        self._show_panel(panel)
    
    def _show_panel(self, panel):
        """Mostrar un panel específico"""
        if self.panel_activo:
            self.panel_activo.pack_forget()
        
        self.panel_activo = panel
        panel.pack(fill=BOTH, expand=True)
    
    def _animate_menus_hide(self):
        """Animar el ocultamiento de header y sidebar"""
        # Animar header hacia arriba
        self._animate_header_hide()
        # Animar sidebar hacia la izquierda  
        self._animate_sidebar_hide()
        # Actualizar el botón de navegación
        self.nav_toggle_btn.configure(text="▶ Mostrar Menús")

    def _animate_header_hide(self):
        """Animar ocultamiento del header hacia arriba"""
        def slide_up(steps_remaining):
            if steps_remaining > 0:
                current_height = self.header.winfo_height()
                new_height = max(0, current_height - 10)
                if new_height > 0:
                    self.after(20, lambda: slide_up(steps_remaining - 1))
                else:
                    self.header.pack_forget()
                    self.header_visible = False
            
        if self.header_visible:
            slide_up(10)

    def _animate_sidebar_hide(self):
        """Animar ocultamiento del sidebar hacia la izquierda"""
        def slide_left(steps_remaining):
            if steps_remaining > 0:
                current_width = self.sidebar.winfo_width()
                new_width = max(0, current_width - 30)
                self.sidebar.config(width=new_width)
                if new_width > 0:
                    self.after(20, lambda: slide_left(steps_remaining - 1))
                else:
                    self.sidebar.pack_forget()
                    self.sidebar_visible = False
            
        if self.sidebar_visible:
            slide_left(8)

    def _toggle_navigation_visibility(self):
        """Alternar visibilidad de header y sidebar con animación"""
        if self.header_visible and self.sidebar_visible:
            # Ocultar menús
            self._animate_menus_hide()
        else:
            # Mostrar menús
            self._animate_menus_show()

    def _animate_menus_show(self):
        """Animar la aparición de header y sidebar"""
        # Mostrar header
        if not self.header_visible:
            # Obtener el contenedor padre correcto
            parent = self.content_container.master
            self.header.pack(fill=X, before=parent.children[list(parent.children.keys())[0]])
            self.header_visible = True
            
        # Mostrar sidebar
        if not self.sidebar_visible:
            # Obtener el contenedor padre correcto (body frame)
            body_parent = self.content_container.master
            self.sidebar.pack(side=LEFT, fill=Y, before=self.content_container)
            self.sidebar.config(width=self.sidebar_width)
            self.sidebar_visible = True
            
        # Actualizar botón
        self.nav_toggle_btn.configure(text="◀ Ocultar Menús")
        
        # Si no estamos en la pantalla de bienvenida, salir del modo bienvenida
        if self.welcome_screen_active and self.panel_activo != self.welcome_frame:
            self.welcome_screen_active = False

    def open_web_auto_modal(self):
        """Abrir modal de automatización web"""
        messagebox.showinfo("Web Automation", "Función de automatización web - En desarrollo")



    # Métodos de utilidad (conservando la lógica de carga de archivos)
    def _get_path(self, relative_path):
        """Obtiene la ruta absoluta de un archivo, compatible con PyInstaller"""
        try: 
            base_path = sys._MEIPASS
        except Exception: 
            base_path = os.path.abspath(os.path.dirname(__file__))
        return os.path.join(base_path, relative_path)
    
    def _cargar_foto_usuario(self):
        """Cargar foto del usuario desde la ruta especificada"""
        try:
            ruta_foto = self.info_usuario.get("ruta_foto", "SIN_FOTO")
            if ruta_foto and ruta_foto != "SIN_FOTO" and os.path.exists(ruta_foto):
                img = Image.open(ruta_foto).resize((72, 72), Image.Resampling.LANCZOS)
                return ImageTk.PhotoImage(img)
            return None
        except Exception as e:
            logging.error(f"Error al cargar la foto del usuario: {e}")
            return None

    def _cargar_iconos(self):
        """Carga todos los iconos necesarios para la interfaz"""
        iconos = {}
        icon_files = {
            "logo1": "logo1.png", 
            "logo2": "logo2.png", 
            "logo3": "logo3.png",
            "usuario": "usuario.png"
        }
        for name, filename in icon_files.items():
            try:
                path = self._get_path(os.path.join("imagenes", filename))
                if name in ["logo1", "logo3"]:
                    size = (110, 110)     # logos header
                elif name == "logo2":
                    size = (140, 140)
                elif name == "usuario":
                    size = (64, 64)       # foto usuario por defecto
                else:
                    size = (32, 32)
                img = Image.open(path).resize(size, Image.Resampling.LANCZOS)
                iconos[name] = ImageTk.PhotoImage(img)
            except Exception as e:
                logging.warning(f"No se encontró el icono o logo '{filename}': {e}")
                iconos[name] = None
        return iconos

    # =========================
    # Helpers UI (KPI y Status)
    # =========================
    # =========================
    # Métodos de utilidad y renderizado
    # =========================
    def _render_kpis(self, df):
        """Actualizar los valores de los KPIs con datos del DataFrame"""
        if df is None or df.empty:
            # Resetear valores cuando no hay datos
            try:
                if hasattr(self, 'kpi_total') and hasattr(self.kpi_total, 'value_lbl'):
                    self.kpi_total.value_lbl.configure(text="0")
                if hasattr(self, 'kpi_rango') and hasattr(self.kpi_rango, 'value_lbl'):
                    self.kpi_rango.value_lbl.configure(text="—")
                if hasattr(self, 'kpi_ultimo') and hasattr(self.kpi_ultimo, 'value_lbl'):
                    self.kpi_ultimo.value_lbl.configure(text="—")
            except Exception as e:
                logging.error(f"Error al resetear KPIs: {e}")
            return

        total = len(df)

        # Rango de fechas: fecha mínima y máxima
        fecha_cols = [
            "Fecha finalizacion (3. Fecha del informe)",
            "Fecha de informe",
            "Fecha de ingreso",
        ]
        fecha_min = None
        fecha_max = None
        
        for c in fecha_cols:
            if c in df.columns:
                try:
                    fechas_validas = pd.to_datetime(df[c], dayfirst=True, errors="coerce").dropna()
                    if not fechas_validas.empty:
                        if fecha_min is None or fechas_validas.min() < fecha_min:
                            fecha_min = fechas_validas.min()
                        if fecha_max is None or fechas_validas.max() > fecha_max:
                            fecha_max = fechas_validas.max()
                except Exception:
                    pass

        # Formatear rango de fechas
        if fecha_min and fecha_max:
            if fecha_min.date() == fecha_max.date():
                rango_txt = fecha_min.strftime("%d/%m/%Y")
            else:
                rango_txt = f"{fecha_min.strftime('%d/%m/%Y')} - {fecha_max.strftime('%d/%m/%Y')}"
        else:
            rango_txt = "—"

        # Última importación (fecha más reciente)
        ultimo_txt = "—" if (fecha_max is None or pd.isna(fecha_max)) else fecha_max.strftime("%d/%m/%Y")

        # Actualizar KPIs con los nuevos valores
        try:
            if hasattr(self, 'kpi_total') and hasattr(self.kpi_total, 'value_lbl'):
                self.kpi_total.value_lbl.configure(text=f"{total:,}".replace(",", "."))
            if hasattr(self, 'kpi_rango') and hasattr(self.kpi_rango, 'value_lbl'):
                self.kpi_rango.value_lbl.configure(text=rango_txt)
            if hasattr(self, 'kpi_ultimo') and hasattr(self.kpi_ultimo, 'value_lbl'):
                self.kpi_ultimo.value_lbl.configure(text=ultimo_txt)
        except Exception as e:
            logging.error(f"Error al actualizar KPIs: {e}")

    def set_status(self, text):
        """Actualiza el texto de la barra de estado"""
        try:
            # Crear una barra de estado temporal si no existe
            if not hasattr(self, 'status_label'):
                self.status_label = ttk.Label(self, text=text, bootstyle="secondary")
                self.status_label.pack(side="bottom", fill="x", padx=5, pady=2)
            else:
                self.status_label.configure(text=text)
        except Exception as e:
            logging.error(f"Error al actualizar status: {e}")
            print(f"[STATUS] {text}")  # Fallback al console

    # =========================
    # Métodos de carga de recursos (conservados)
    # =========================
            pass

    # =========================
    # Navegación (métodos obsoletos removidos - usar los métodos nuevos)
    # =========================




    # ---------- Helpers Dashboard ----------

    def _clear_filters(self):
        for k in self.db_filters:
            self.db_filters[k].set("")
        self._refresh_dashboard()

    def _refresh_dashboard(self):
        try:
            self.set_status("Actualizando dashboard…")
            self.cargar_dashboard()
        finally:
            self.set_status("Dashboard actualizado.")

    def _get_filtered_df(self, df):
        dff = df.copy()
        fd = self.db_filters["fecha_desde"].get().strip()
        fh = self.db_filters["fecha_hasta"].get().strip()
        if fd:
            d0 = pd.to_datetime(fd, dayfirst=True, errors="coerce")
            if pd.notna(d0):
                dff = dff[dff["_fecha_informe"] >= d0]
        if fh:
            d1 = pd.to_datetime(fh, dayfirst=True, errors="coerce")
            if pd.notna(d1):
                dff = dff[dff["_fecha_informe"] <= d1]

        srv = self.db_filters["servicio"].get().strip()
        if srv:
            dff = dff[dff.get("Servicio", "").astype(str).eq(srv)]
        mal = self.db_filters["malignidad"].get().strip()
        if mal:
            dff = dff[dff.get("Malignidad", "").astype(str).str.upper().eq(mal)]
        rsp = self.db_filters["responsable"].get().strip()
        if rsp:
            dff = dff[dff.get("Usuario finalizacion", "").astype(str).eq(rsp)]
        return dff

    def _clear_dash_area(self):
        # Desmonta los canvases previos para liberar memoria
        for cv in getattr(self, "_dash_canvases", []):
            try:
                cv.get_tk_widget().destroy()
            except Exception:
                pass
        self._dash_canvases = []

        # Limpia frames hijos en cada pestaña
        for tab in [self.tab_overview, self.tab_biomarkers, self.tab_times, self.tab_quality, self.tab_compare]:
            for child in tab.grid_slaves():
                child.destroy()

    def _chart_in(self, tab, row, col, render_fn, title, dff):
        card = ttk.Frame(tab, padding=8, relief="solid", borderwidth=1)
        card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
        
        # Configurar responsive grid
        tab.grid_rowconfigure(row, weight=1)
        tab.grid_columnconfigure(col, weight=1)

        header = ttk.Frame(card)
        header.pack(fill="x", pady=(0, 6))
        ttk.Label(header, text=title, font=("Segoe UI", 11, "bold")).pack(side="left")
        ttk.Button(header, text="⛶", 
                  command=lambda: self._open_fullscreen_figure(render_fn, title, dff),
                  bootstyle="outline").pack(side="right")

        try:
            fig = render_fn()
            if fig is None:
                ttk.Label(card, text="(sin datos)", font=("Segoe UI", 10)).pack(padx=10, pady=20)
                return
            
            # Ajustar tamaño de figura para mejor responsive
            fig.set_size_inches(4.5, 3.2)
            fig.set_dpi(90)
            
            canvas = FigureCanvasTkAgg(fig, master=card)
            canvas.draw()
            widget = canvas.get_tk_widget()
            widget.pack(fill="both", expand=True, padx=4, pady=4)
            widget.bind("<Double-Button-1>", lambda e: self._open_fullscreen_figure(render_fn, title, dff))
            self._dash_canvases.append(canvas)
        except Exception as e:
            ttk.Label(card, text=f"Error: {e}", font=("Segoe UI", 9), bootstyle="danger").pack(padx=10, pady=10)

    def _toggle_db_sidebar(self):
        # Muestra/oculta el sidebar, ajusta texto del botón y grid
        if self.db_sidebar_collapsed:
            self.db_sidebar.grid(row=0, column=0, sticky="ns", padx=(0, 12), pady=6)
            self.btn_toggle_sidebar.configure(text="✕ Ocultar filtros")
        else:
            self.db_sidebar.grid_forget()
            self.btn_toggle_sidebar.configure(text="≡ Mostrar filtros")
        self.db_sidebar_collapsed = not self.db_sidebar_collapsed

    def _open_filters_sheet(self):
        # Modal de filtros (para no robar ancho)
        top = tk.Toplevel(self)
        top.title("Filtros")
        top.geometry("420x420")
        top.grab_set()
        top.transient(self)
        
        wrap = ttk.Frame(top, padding=12)
        wrap.pack(fill="both", expand=True)

        def row(lbl, widget):
            r = ttk.Frame(wrap)
            r.pack(fill="x", pady=6)
            ttk.Label(r, text=lbl, width=20, anchor="w").pack(side="left")
            widget.pack(in_=r, side="left", fill="x", expand=True)

        e1 = ttk.Entry(wrap, textvariable=self.db_filters["fecha_desde"])
        e2 = ttk.Entry(wrap, textvariable=self.db_filters["fecha_hasta"])
        
        # Verificar si los componentes existen antes de obtener sus valores
        servicio_vals = []
        if self.cmb_servicio is not None:
            try:
                servicio_vals = list(self.cmb_servicio.cget("values"))
            except:
                servicio_vals = []
        
        malig_vals = ["", "PRESENTE", "AUSENTE"]
        if self.cmb_malig is not None:
            try:
                malig_vals = list(self.cmb_malig.cget("values"))
            except:
                malig_vals = ["", "PRESENTE", "AUSENTE"]
        
        resp_vals = []
        if self.cmb_resp is not None:
            try:
                resp_vals = list(self.cmb_resp.cget("values"))
            except:
                resp_vals = []
        
        cb1 = ttk.Combobox(wrap, values=servicio_vals, textvariable=self.db_filters["servicio"])
        cb2 = ttk.Combobox(wrap, values=malig_vals, textvariable=self.db_filters["malignidad"])
        cb3 = ttk.Combobox(wrap, values=resp_vals, textvariable=self.db_filters["responsable"])

        row("Fecha desde (dd/mm/aaaa)", e1)
        row("Fecha hasta (dd/mm/aaaa)", e2)
        row("Servicio", cb1)
        row("Malignidad", cb2)
        row("Responsable", cb3)

        btns = ttk.Frame(wrap)
        btns.pack(fill="x", pady=(10,0))
        ttk.Button(btns, text="Aplicar", command=lambda:(self._refresh_dashboard(), top.destroy())).pack(side="left", expand=True, fill="x", padx=(0,6))
        ttk.Button(btns, text="Limpiar", command=self._clear_filters).pack(side="left", expand=True, fill="x", padx=(6,0))

    def _open_fullscreen_figure(self, render_fn, title, dff):
        # Ventana a pantalla completa con inspector lateral
        fs = tk.Toplevel(self)
        fs.title(title)
        try:
            fs.state('zoomed')
        except Exception:
            pass
        fs.grid_rowconfigure(0, weight=1)
        fs.grid_columnconfigure(0, weight=1)
        fs.grid_columnconfigure(1, weight=0)

        # Área de gráfico
        graph_area = ttk.Frame(fs, padding=10)
        graph_area.grid(row=0, column=0, sticky="nsew", padx=(10,6), pady=10)
        fig = render_fn()
        if fig is None:
            ttk.Label(graph_area, text="(sin datos)").pack(padx=12, pady=12)
        else:
            canv = FigureCanvasTkAgg(fig, master=graph_area)
            canv.draw()
            canv.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)

        # Inspector lateral
        insp = ttk.Frame(fs, padding=10, width=300)
        insp.grid(row=0, column=1, sticky="ns", padx=(6,10), pady=10)
        ttk.Label(insp, text="Inspector", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0,6))
        self._build_inspector(insp, title, dff)

        # Barra superior simple (cerrar)
        topbar = ttk.Frame(graph_area)
        topbar.pack(fill="x", pady=(0,10))
        ttk.Label(topbar, text=title, font=("Segoe UI", 14, "bold")).pack(side="left")
        ttk.Button(topbar, text="Cerrar", command=fs.destroy).pack(side="right")

    def _build_inspector(self, parent, title, dff):
        # Datos generales
        n = len(dff)
        fmin = pd.to_datetime(dff.get("_fecha_informe"), errors="coerce").min()
        fmax = pd.to_datetime(dff.get("_fecha_informe"), errors="coerce").max()
        rng = f"{fmin:%d/%m/%Y} – {fmax:%d/%m/%Y}" if pd.notna(fmin) and pd.notna(fmax) else "—"

        def row(k, v):
            r = ttk.Frame(parent)
            r.pack(fill="x", padx=12, pady=4)
            ttk.Label(r, text=k).pack(side="left")
            ttk.Label(r, text=v).pack(side="right")

        row("Registros filtrados", f"{n:,}".replace(",", "."))
        row("Rango de fechas", rng)

        # Secciones condicionales útiles
        if "Malignidad" in dff.columns:
            ser = dff["Malignidad"].astype(str).str.upper().value_counts()
            box = ttk.Frame(parent, padding=8, relief="solid", borderwidth=1); box.pack(fill="x", padx=12, pady=(10,4))
            ttk.Label(box, text="Malignidad", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=10, pady=(8,2))
            for k,v in ser.items():
                rowtxt = ttk.Frame(box); rowtxt.pack(fill="x", padx=10, pady=2)
                ttk.Label(rowtxt, text=f"{k}").pack(side="left")
                ttk.Label(rowtxt, text=str(v)).pack(side="right")

        if "Organo (1. Muestra enviada a patología)" in dff.columns:
            top_org = dff["Organo (1. Muestra enviada a patología)"].astype(str).replace({"": "No especificado"}).value_counts().head(8)
        elif "IHQ_ORGANO" in dff.columns:
            top_org = dff["IHQ_ORGANO"].astype(str).replace({"": "No especificado"}).value_counts().head(8)
        else:
            top_org = None

        if top_org is not None and not top_org.empty:
            box2 = ttk.Frame(parent, padding=8, relief="solid", borderwidth=1); box2.pack(fill="x", padx=12, pady=(10,12))
            ttk.Label(box2, text="Top Órganos", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=10, pady=(8,2))
            for k,v in top_org.items():
                rowtxt = ttk.Frame(box2); rowtxt.pack(fill="x", padx=10, pady=2)
                ttk.Label(rowtxt, text=f"{k}").pack(side="left")
                ttk.Label(rowtxt, text=str(v)).pack(side="right")

    # ---------- Renderers: Overview ----------

    def _g_line_informes_por_mes(self, df):
        if df.empty or df["_fecha_informe"].isna().all():
            return None
        ser = df.dropna(subset=["_fecha_informe"]).set_index("_fecha_informe").resample("MS").size()
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.plot(ser.index, ser.values, marker="o")
        ax.set_title("Informes por mes")
        ax.set_xlabel("Mes")
        ax.set_ylabel("Conteo")
        fig.tight_layout()
        return fig

    def _g_pie_malignidad(self, df):
        if "Malignidad" not in df.columns or df.empty:
            return None
        ser = df["Malignidad"].astype(str).str.upper().replace({"": "DESCONOCIDO"}).value_counts()
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.pie(ser.values, labels=ser.index, autopct="%1.1f%%", startangle=90)
        ax.set_title("Distribución de Malignidad")
        fig.tight_layout()
        return fig

    def _g_bar_top_servicio(self, df, top=12):
        if "Servicio" not in df.columns or df.empty:
            return None
        ser = df["Servicio"].astype(str).value_counts().head(top)
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title(f"Top Servicios (n={ser.sum()})")
        ax.set_ylabel("Informes")
        ax.tick_params(axis="x", rotation=30, labelsize=8)
        fig.tight_layout()
        return fig

    def _g_bar_top_organo(self, df, top=12):
        # soporta tanto columna Excel como IHQ_ORGANO
        col = "Organo (1. Muestra enviada a patología)" if "Organo (1. Muestra enviada a patología)" in df.columns else ("IHQ_ORGANO" if "IHQ_ORGANO" in df.columns else None)
        if not col: return None
        ser = df[col].astype(str).replace({"": "No especificado"}).value_counts().head(top)
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title("Top Órganos")
        ax.set_ylabel("Informes")
        ax.tick_params(axis="x", rotation=30, labelsize=8)
        fig.tight_layout()
        return fig

    # ---------- Renderers: Biomarcadores ----------

    def _g_hist_ki67(self, df):
        col = "IHQ_KI-67" if "IHQ_KI-67" in df.columns else None
        if not col: return None
        
        # Limpiar y convertir los datos
        raw_data = df[col].astype(str)  # Convertir todo a string primero
        # Remover espacios y reemplazar cadenas vacías con NaN
        clean_data = raw_data.str.strip().replace('', pd.NA)
        # Remover el símbolo % y convertir a numérico
        numeric_data = clean_data.str.replace('%', '', regex=False)
        s = pd.to_numeric(numeric_data, errors="coerce").dropna()
        
        if s.empty: 
            print("DEBUG: No hay datos numéricos válidos para Ki-67")
            return None
            
        print(f"DEBUG: Ki-67 datos válidos encontrados: {len(s)} valores, rango: {s.min()}-{s.max()}")
        
        fig = Figure(figsize=(5.6, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        ax.hist(s.values, bins=min(12, len(s.unique())))  # Ajustar bins si hay pocos datos únicos
        ax.set_title("Ki-67 (%)")
        ax.set_xlabel("%")
        ax.set_ylabel("Frecuencia")
        fig.tight_layout()
        return fig

    def _g_bar_her2(self, df):
        col = "IHQ_HER2" if "IHQ_HER2" in df.columns else None
        if not col: return None
        order = ["0", "1+", "2+", "3+", "NEGATIVO", "POSITIVO"]
        ser = df[col].astype(str).str.upper().value_counts()
        ser = ser.reindex(order, fill_value=0) if any(k in ser.index for k in order) else ser
        if ser.sum() == 0: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title("HER2 (score)")
        ax.set_ylabel("Informes")
        fig.tight_layout()
        return fig

    def _g_bar_re_rp(self, df):
        cols = [c for c in ["IHQ_RECEPTOR_ESTROGENO", "IHQ_RECEPTOR_PROGESTAGENOS"] if c in df.columns]
        if not cols: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        data = []
        labels = []
        for c in cols:
            ser = df[c].astype(str).str.upper().replace({"": "ND"}).value_counts()
            data.append(ser)
            labels.append(c.replace("IHQ_", ""))
        # Normaliza categorías
        cats = sorted(set().union(*[d.index for d in data]))
        mat = np.array([[d.get(k, 0) for k in cats] for d in data])
        for i, row in enumerate(mat):
            ax.bar(np.arange(len(cats))+i*0.35, row, width=0.35, label=labels[i])
        ax.set_xticks(np.arange(len(cats))+0.35/2)
        ax.set_xticklabels(cats, rotation=0)
        ax.set_title("RE / RP (estado)")
        ax.legend()
        fig.tight_layout()
        return fig

    def _g_bar_pdl1(self, df):
        # intenta TPS o CPS
        for col in ["IHQ_PDL-1", "IHQ_PDL1_TPS", "IHQ_PDL1_CPS"]:
            if col in df.columns:
                s = df[col].astype(str)
                break
        else:
            return None
        ser = s.replace({"": "ND"}).value_counts()
        if ser.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title("PD-L1")
        ax.set_ylabel("Informes")
        ax.tick_params(axis="x", rotation=0)
        fig.tight_layout()
        return fig

    # ---------- Renderers: Tiempos ----------

    def _g_box_tiempo_proceso(self, df):
        f_ing = pd.to_datetime(df.get("Fecha de ingreso (2. Fecha de la muestra)", ""), dayfirst=True, errors="coerce")
        f_inf = pd.to_datetime(df.get("Fecha finalizacion (3. Fecha del informe)", df.get("Fecha de informe", "")), dayfirst=True, errors="coerce")
        dias = (f_inf - f_ing).dt.days.dropna()
        if dias.empty: return None
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.boxplot(dias.values, vert=True)
        ax.set_title("Tiempo de proceso (días)")
        fig.tight_layout()
        return fig

    def _g_line_throughput_semana(self, df):
        if df.empty or df["_fecha_informe"].isna().all(): return None
        ser = df.dropna(subset=["_fecha_informe"]).set_index("_fecha_informe").resample("W-MON").size()
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.plot(ser.index, ser.values, marker="o")
        ax.set_title("Throughput semanal")
        ax.set_xlabel("Semana")
        ax.set_ylabel("Informes")
        fig.tight_layout()
        return fig

    def _g_scatter_edad_ki67(self, df):
        if "Edad" not in df.columns: return None
        x = pd.to_numeric(df["Edad"], errors="coerce")
        
        # Limpiar los datos Ki-67 igual que en _g_hist_ki67
        if "IHQ_KI-67" in df.columns:
            raw_ki67 = df["IHQ_KI-67"].astype(str).str.strip().replace('', pd.NA)
            clean_ki67 = raw_ki67.str.replace('%', '', regex=False)
            y = pd.to_numeric(clean_ki67, errors="coerce")
        else:
            y = pd.Series(np.nan, index=df.index)
            
        m = x.notna() & y.notna()
        if not m.any(): 
            print("DEBUG: No hay datos válidos para scatter Edad vs Ki-67")
            return None
            
        print(f"DEBUG: Scatter Edad vs Ki-67 - {m.sum()} puntos válidos")
        
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.scatter(x[m], y[m], alpha=0.6)
        ax.set_title("Edad vs Ki-67")
        ax.set_xlabel("Edad")
        ax.set_ylabel("Ki-67 (%)")
        fig.tight_layout()
        return fig

    # ---------- Renderers: Calidad ----------

    def _g_bar_missingness(self, df):
        cols = [
            "Servicio", "Malignidad", "Usuario finalizacion",
            "Organo (1. Muestra enviada a patología)", "IHQ_HER2", "IHQ_KI-67"
        ]
        present = [c for c in cols if c in df.columns]
        if not present: return None
        miss = df[present].isna().mean().sort_values(ascending=False)
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(miss.index, (miss.values*100.0))
        ax.set_title("Campos vacíos (%)")
        ax.set_ylabel("% vacío")
        ax.tick_params(axis="x", rotation=25)
        fig.tight_layout()
        return fig

    def _g_bar_top_responsables(self, df, top=10):
        col = "Usuario finalizacion"
        if col not in df.columns: return None
        ser = df[col].astype(str).value_counts().head(top)
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(ser.index, ser.values)
        ax.set_title("Productividad por responsable (Top)")
        ax.set_ylabel("Informes")
        ax.tick_params(axis="x", rotation=25)
        fig.tight_layout()
        return fig

    def _g_bar_largos_texto(self, df):
        col = "Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)"
        if col not in df.columns: return None
        s = df[col].astype(str).str.len()
        bins = [0, 50, 150, 300, 600, 1200, np.inf]
        ser = pd.cut(s, bins=bins, labels=["<50", "50–150", "150–300", "300–600", "600–1200", "1200+"], include_lowest=True).value_counts().sort_index()
        fig = Figure(figsize=(5.6, 3.2), dpi=100); ax = fig.add_subplot(111)
        ax.bar(ser.index.astype(str), ser.values)
        ax.set_title("Longitud del diagnóstico (bins)")
        ax.set_ylabel("Informes")
        fig.tight_layout()
        return fig

    # ---------- Comparador parametrizable ----------

    def _build_comparator(self, tab, df):
        # Controles
        ctrl = ttk.Frame(tab)
        ctrl.grid(row=0, column=0, columnspan=2, padx=10, pady=(10,0), sticky="ew")

        dims = [c for c in ["Servicio", "Usuario finalizacion", "Malignidad", "Organo (1. Muestra enviada a patología)"] if c in df.columns]
        mets = [c for c in ["IHQ_KI-67"] if c in df.columns]  # se pueden añadir más numéricas

        self._compare_controls["dim"] = tk.StringVar(value=dims[0] if dims else "")
        self._compare_controls["agg"] = tk.StringVar(value="conteo")
        self._compare_controls["met"] = tk.StringVar(value=mets[0] if mets else "")

        row = ttk.Frame(ctrl, padding=10, relief="solid", borderwidth=1)
        row.pack(fill="x", padx=4, pady=4)
        ttk.Label(row, text="Dimensión:").pack(side="left", padx=6)
        ttk.Combobox(row, values=dims or [""], textvariable=self._compare_controls["dim"]).pack(side="left", padx=6)
        ttk.Label(row, text="Agregador:").pack(side="left", padx=6)
        ttk.Combobox(row, values=["conteo", "promedio"], textvariable=self._compare_controls["agg"]).pack(side="left", padx=6)
        ttk.Label(row, text="Métrica:").pack(side="left", padx=6)
        ttk.Combobox(row, values=mets or [""], textvariable=self._compare_controls["met"]).pack(side="left", padx=6)
        ttk.Button(row, text="Aplicar", command=lambda: self._chart_in(tab, 1, 0, lambda: self._g_compare(df), "Comparación de Datos", df)).pack(side="left", padx=10)

        # Gráfico inicial
        self._chart_in(tab, 1, 0, lambda: self._g_compare(df), "Comparación de Datos", df)

    def _g_compare(self, df):
        dim = self._compare_controls["dim"].get()
        agg = self._compare_controls["agg"].get()
        met = self._compare_controls["met"].get()
        if not dim or df.empty: return None

        fig = Figure(figsize=(11.6, 3.2), dpi=100); ax = fig.add_subplot(111)

        if agg == "conteo":
            ser = df[dim].astype(str).value_counts()
            ax.bar(ser.index, ser.values)
            ax.set_title(f"Conteo por {dim}")
            ax.tick_params(axis="x", rotation=25)
        else:
            if not met or met not in df.columns:
                return None
            s = pd.to_numeric(df[met], errors="coerce")
            grp = df.assign(_metric=s).groupby(dim)["_metric"].mean().dropna()
            ax.bar(grp.index, grp.values)
            ax.set_title(f"Promedio de {met} por {dim}")
            ax.tick_params(axis="x", rotation=25)

        fig.tight_layout()
        return fig

    # =========================
    # Funcionalidad
    # =========================
    def log_to_widget(self, message):
        self.log_textbox.configure(state="normal")
        self.log_textbox.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_textbox.see("end")
        self.log_textbox.configure(state="disabled")

    def select_files(self):
        self.pdf_files = filedialog.askopenfilenames(title="Seleccione archivos PDF", filetypes=[("PDF", "*.pdf")])
        if self.pdf_files:
            self.log_to_widget(f"Seleccionados {len(self.pdf_files)} archivos.")
            self.start_button.configure(state="normal")
            self.set_status(f"{len(self.pdf_files)} archivos listos.")
        else:
            self.log_to_widget("Selección cancelada.")
            self.start_button.configure(state="disabled")
            self.set_status("Selección cancelada.")

    def start_processing(self):
        if not self.pdf_files:
            messagebox.showwarning("Advertencia", "Por favor, seleccione archivos PDF primero.")
            return

        self.start_button.configure(state="disabled")
        self.select_files_button.configure(state="disabled")
        self.log_to_widget("=" * 50)
        self.log_to_widget("INICIANDO PROCESAMIENTO...")
        self.set_status("Procesando… esto puede tardar según el tamaño de los PDFs.")

        threading.Thread(target=self.processing_thread, daemon=True).start()

    def processing_thread(self):
        try:
            output_dir = os.path.dirname(self.pdf_files[0])
            num_records = procesador_ihq_biomarcadores.process_ihq_paths(self.pdf_files, output_dir)
            self.log_to_widget(f"PROCESO COMPLETADO. Se guardaron {num_records} registros en la base de datos.")
            messagebox.showinfo("Éxito", f"Proceso finalizado. Se guardaron {num_records} registros.")
            self.set_status("Completado ✔  |  " + datetime.now().strftime("%d/%m/%Y %H:%M"))
        except Exception as e:
            self.log_to_widget(f"ERROR: {e}")
            messagebox.showerror("Error", f"Ocurrió un error durante el procesamiento:\n{e}")
            self.set_status("Error en el procesamiento.")
        finally:
            self.start_button.configure(state="normal")
            self.select_files_button.configure(state="normal")

    def refresh_data_and_table(self):
        try:
            from database_manager import init_db, get_all_records_as_dataframe
            init_db()
            self.master_df = get_all_records_as_dataframe()
            self._populate_treeview(self.master_df)
            self._render_kpis(self.master_df)
        except Exception as e:
            messagebox.showerror("Error de Base de Datos", f"No se pudieron cargar los datos: {e}")
            self.set_status("Error al cargar datos.")

    def _populate_treeview(self, df_to_display):
        # Verificar que el tree exista antes de usar
        if self.tree is None:
            return
            
        for item in self.tree.get_children():
            self.tree.delete(item)

        if df_to_display.empty:
            return

        cols_to_show = [
            "N. peticion (0. Numero de biopsia)",
            "Primer nombre",
            "Primer apellido",
            "Fecha finalizacion (3. Fecha del informe)",
            "Malignidad",
            "Organo (1. Muestra enviada a patología)",
        ]
        df_display = df_to_display[[c for c in cols_to_show if c in df_to_display.columns]].copy()

        self.tree["columns"] = list(df_display.columns)
        for col in df_display.columns:
            header = col.split("(")[0].strip()
            # Ordenamiento al clic
            self.tree.heading(col, text=header, command=lambda c=col: self._sort_treeview(c, False))
            # Auto-ancho (acorde a datos y encabezado, con límites)
            try:
                max_len = max(df_display[col].astype(str).str.len().max(), len(header))
            except Exception:
                max_len = len(header)
            width = max(120, min(280, int(max_len * 7)))
            self.tree.column(col, width=width, anchor="w", stretch=True)

        # Filas cebra
        self.tree.tag_configure("oddrow", background="#2a2d2e")
        self.tree.tag_configure("evenrow", background="#232629")

        for idx, (_, row) in enumerate(df_display.iterrows()):
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert("", "end", values=list(row), iid=idx, tags=(tag,))

        # Actualizar KPIs en base a lo mostrado
        try:
            self._render_kpis(df_display)
        except Exception:
            pass

    def _sort_treeview(self, col, reverse):
        items = self.tree.get_children("")
        data = [(self.tree.set(it, col), it) for it in items]

        from datetime import datetime as _dt

        def _key(v):
            s = str(v).strip()
            # número
            try:
                return float(s.replace(",", "."))
            except Exception:
                pass
            # fecha
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return _dt.strptime(s, fmt)
                except Exception:
                    pass
            return s.lower()

        data.sort(key=lambda x: _key(x[0]), reverse=reverse)

        for idx, (_, item) in enumerate(data):
            self.tree.move(item, "", idx)

        self.tree.heading(col, command=lambda: self._sort_treeview(col, not reverse))

    def filter_tabla(self, *args):
        query = self.search_var.get().lower()
        if not query:
            self._populate_treeview(self.master_df)
            return

        df = self.master_df.copy()
        if df.empty:
            self._populate_treeview(df)
            return
            
        search_cols = ["N. peticion (0. Numero de biopsia)", "Primer nombre", "Primer apellido"]
        for col in search_cols:
            if col in df.columns:
                df[col] = df[col].astype(str)
        
        # Inicializar máscara con todos False
        mask = pd.Series([False] * len(df), index=df.index)
        
        if search_cols[0] in df.columns:
            mask |= df[search_cols[0]].str.lower().str.contains(query, na=False)
        if search_cols[1] in df.columns:
            mask |= df[search_cols[1]].str.lower().str.contains(query, na=False)
        if search_cols[2] in df.columns:
            mask |= df[search_cols[2]].str.lower().str.contains(query, na=False)
        
        self._populate_treeview(df[mask])

    def mostrar_detalle_registro(self, event):
        selected_item = self.tree.focus()
        if not selected_item:
            return

        item_index = int(selected_item)
        if item_index not in self.master_df.index:
            return
        record = self.master_df.loc[item_index]

        self.detail_textbox.configure(state="normal")
        self.detail_textbox.delete("1.0", "end")

        details_text = ""
        for key, value in record.items():
            if pd.notna(value) and str(value).strip():
                details_text += f"{key.split('(')[0].strip()}:\n{value}\n{'-'*30}\n"

        self.detail_textbox.insert("1.0", details_text)
        self.detail_textbox.configure(state="disabled")

    def cargar_dashboard(self):
        # 1) Preparar DF y combos de filtros
        df = self.master_df.copy()
        if df is None or df.empty:
            self._render_kpis(df)
            self._clear_dash_area()
            return

        # Normaliza fechas (varias columnas posibles)
        df["_fecha_informe"] = pd.to_datetime(
            df.get("Fecha finalizacion (3. Fecha del informe)", df.get("Fecha de informe", df.get("Fecha de ingreso", ""))),
            dayfirst=True, errors="coerce"
        )

        # Llenar combos dinámicos (servicios / responsables)
        srv_vals = sorted([s for s in df.get("Servicio", pd.Series(dtype=str)).dropna().astype(str).unique() if s.strip()])
        rsp_vals = sorted([s for s in df.get("Usuario finalizacion", pd.Series(dtype=str)).dropna().astype(str).unique() if s.strip()])
        
        # Solo configurar si los componentes existen
        if self.cmb_servicio is not None:
            try:
                self.cmb_servicio.configure(values=[""] + srv_vals)
            except:
                pass
        
        if self.cmb_resp is not None:
            try:
                self.cmb_resp.configure(values=[""] + rsp_vals)
            except:
                pass

        # 2) Render de KPIs básicos
        self._render_kpis(df)

        # 3) Limpiar canvases anteriores y pintar
        self._clear_dash_area()

        # Filtros iniciales (los que estén llenos)
        dff = self._get_filtered_df(df)

        # 4) PINTAR: OVERVIEW (4 gráficos)
        self._chart_in(self.tab_overview, 0, 0, lambda: self._g_line_informes_por_mes(dff), "Informes por mes", dff)
        self._chart_in(self.tab_overview, 0, 1, lambda: self._g_pie_malignidad(dff), "Distribución de Malignidad", dff)
        self._chart_in(self.tab_overview, 1, 0, lambda: self._g_bar_top_servicio(dff), "Top Servicios", dff)
        self._chart_in(self.tab_overview, 1, 1, lambda: self._g_bar_top_organo(dff), "Top Órganos", dff)

        # 5) PINTAR: BIOMARCADORES
        self._chart_in(self.tab_biomarkers, 0, 0, lambda: self._g_hist_ki67(dff), "Ki-67 (%)", dff)
        self._chart_in(self.tab_biomarkers, 0, 1, lambda: self._g_bar_her2(dff), "HER2 (score)", dff)
        self._chart_in(self.tab_biomarkers, 1, 0, lambda: self._g_bar_re_rp(dff), "RE / RP (estado)", dff)
        self._chart_in(self.tab_biomarkers, 1, 1, lambda: self._g_bar_pdl1(dff), "PD-L1", dff)

        # 6) PINTAR: TIEMPOS
        self._chart_in(self.tab_times, 0, 0, lambda: self._g_box_tiempo_proceso(dff), "Tiempo de proceso (días)", dff)
        self._chart_in(self.tab_times, 0, 1, lambda: self._g_line_throughput_semana(dff), "Throughput semanal", dff)
        self._chart_in(self.tab_times, 1, 0, lambda: self._g_scatter_edad_ki67(dff), "Edad vs Ki-67", dff)

        # 7) PINTAR: CALIDAD
        self._chart_in(self.tab_quality, 0, 0, lambda: self._g_bar_missingness(dff), "Campos vacíos (%)", dff)
        self._chart_in(self.tab_quality, 0, 1, lambda: self._g_bar_top_responsables(dff), "Productividad por responsable", dff)
        self._chart_in(self.tab_quality, 1, 0, lambda: self._g_bar_largos_texto(dff), "Longitud del diagnóstico", dff)

        # 8) PINTAR: COMPARADOR
        self._build_comparator(self.tab_compare, dff)


    # =========================
    # Estilo de tabla (Treeview) - MÉTODO ACTUALIZADO
    # =========================
    def setup_treeview_style(self):
        style = ttk_std.Style()
        style.theme_use("clam")  # look & feel moderno y estable en Windows

        # Cuerpo de la tabla
        style.configure(
            "Custom.Treeview",
            background="#1f2124",
            fieldbackground="#1f2124",
            foreground="#e9ecef",
            rowheight=30,
            borderwidth=0,
        )
        style.map(
            "Custom.Treeview",
            background=[("selected", "#2b6cb0")],
            foreground=[("selected", "white")],
        )

        # Encabezados
        style.configure(
            "Custom.Treeview.Heading",
            background="#262a2e",
            foreground="#e9ecef",
            font=("Segoe UI", 11, "bold"),
            relief="flat",
        )
        style.map("Custom.Treeview.Heading", background=[("active", "#2d3136")])
        return style
    # ---------- Automatización Web (modal + ejecución) ----------

    def open_web_auto_modal(self):
        top = tk.Toplevel(self)
        top.title("Automatizar Entrega de Resultados")
        top.geometry("460x360")
        top.grab_set()

        top.transient(self)
        try:
            top.lift(); top.focus_force()
        except Exception:
            pass
        
        # Campos
        frm = ttk.Frame(top, padding=12, relief="solid", borderwidth=1)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        # Usuario / Clave
        ttk.Label(frm, text="Usuario").grid(row=0, column=0, padx=10, pady=(12,6), sticky="w")
        user_var = tk.StringVar(value="12345")
        ttk.Entry(frm, textvariable=user_var).grid(row=0, column=1, padx=10, pady=(12,6), sticky="ew")

        ttk.Label(frm, text="Contraseña").grid(row=1, column=0, padx=10, pady=6, sticky="w")
        pass_var = tk.StringVar(value="CONSULTA1")
        ttk.Entry(frm, textvariable=pass_var, show="•").grid(row=1, column=1, padx=10, pady=6, sticky="ew")

        # Criterio
        ttk.Label(frm, text="Buscar por").grid(row=2, column=0, padx=10, pady=6, sticky="w")
        criterio_var = tk.StringVar(value="Fecha de Ingreso")
        ttk.Combobox(frm, values=["Fecha de Ingreso", "Fecha de Finalizacion", "Rango de Peticion", "Datos del Paciente"], textvariable=criterio_var).grid(row=2, column=1, padx=10, pady=6, sticky="ew")

        # Fechas
        fi_var = tk.StringVar(value="")
        ff_var = tk.StringVar(value="")

        def pick_fi():
            sel = CalendarioInteligente.seleccionar_fecha(parent=top, locale='es_CO', codigo_pais_festivos='CO')
            if sel:
                fi_var.set(sel.strftime("%d/%m/%Y"))
            # RE-ADQUIRIR MODAL Y TRAER AL FRENTE
            try:
                top.deiconify()
                # truco para traer al frente en Windows
                top.attributes("-topmost", True); top.attributes("-topmost", False)
                top.lift(); top.focus_force(); top.grab_set()
            except Exception:
                pass

        def pick_ff():
            sel = CalendarioInteligente.seleccionar_fecha(parent=top, locale='es_CO', codigo_pais_festivos='CO')
            if sel:
                ff_var.set(sel.strftime("%d/%m/%Y"))
            # RE-ADQUIRIR MODAL Y TRAER AL FRENTE
            try:
                top.deiconify()
                top.attributes("-topmost", True); top.attributes("-topmost", False)
                top.lift(); top.focus_force(); top.grab_set()
            except Exception:
                pass

        ttk.Label(frm, text="Fecha inicial").grid(row=3, column=0, padx=10, pady=6, sticky="w")
        row_fi = ttk.Frame(frm); row_fi.grid(row=3, column=1, padx=10, pady=6, sticky="ew")
        ttk.Entry(row_fi, textvariable=fi_var).pack(side="left", fill="x", expand=True, padx=(0,6))
        ttk.Button(row_fi, text="Elegir…", width=10, command=pick_fi).pack(side="left")

        ttk.Label(frm, text="Fecha final").grid(row=4, column=0, padx=10, pady=6, sticky="w")
        row_ff = ttk.Frame(frm); row_ff.grid(row=4, column=1, padx=10, pady=6, sticky="ew")
        ttk.Entry(row_ff, textvariable=ff_var).pack(side="left", fill="x", expand=True, padx=(0,6))
        ttk.Button(row_ff, text="Elegir…", width=10, command=pick_ff).pack(side="left")

        # Botones
        btns = ttk.Frame(frm); btns.grid(row=5, column=0, columnspan=2, pady=(12,8), sticky="ew")
        ttk.Button(btns, text="Cancelar", command=top.destroy).pack(side="right", padx=6)
        def go():
            top.destroy()
            self._start_web_automation(
                fi_var.get().strip(), ff_var.get().strip(),
                user_var.get().strip(), pass_var.get().strip(),
                criterio_var.get().strip()
            )
        ttk.Button(btns, text="Iniciar", command=go).pack(side="right", padx=6)

        # grid conf
        frm.grid_columnconfigure(1, weight=1)

    def _start_web_automation(self, fi, ff, user, pwd, criterio):
        if not fi or not ff:
            messagebox.showwarning("Fechas requeridas", "Debe seleccionar fecha inicial y final.")
            return
        self.set_status("Automatizando Entrega de resultados…")
        t = threading.Thread(target=self._run_web_automation, args=(fi, ff, user, pwd, criterio), daemon=True)
        t.start()

    def _run_web_automation(self, fi, ff, user, pwd, criterio):
        try:
            ok = automatizar_entrega_resultados(
                fecha_inicial_ddmmaa=fi,
                fecha_final_ddmmaa=ff,
                cred=Credenciales(usuario=user, clave=pwd),
                criterio=criterio,
                headless=False,
                log_cb=self._log_auto
            )
            if ok:
                self.set_status("Consulta web completada. Revise resultados en el navegador.")
                messagebox.showinfo("Automatización", "Consulta completada en el portal.")
            else:
                self.set_status("Automatización: sin resultado.")
        except Exception as e:
            self.set_status(f"Error en automatización: {e}")
            messagebox.showerror("Automatización", f"Ocurrió un error:\n{e}")

    def _log_auto(self, msg: str):
        try:
            # Si está visible el textbox de logs de Procesar, úsalo; si no, status.
            if hasattr(self, "log_textbox") and str(self.log_textbox.winfo_exists()) == "1":
                self.log_textbox.configure(state="normal")
                self.log_textbox.insert("end", f"[AUTO] {msg}\n")
                self.log_textbox.configure(state="disabled")
                self.log_textbox.see("end")
            else:
                self.set_status(msg)
        except Exception:
            self.set_status(msg)

    # =========================
    # Tema claro/oscuro
    # =========================
    # =========================
    # Métodos para el panel de procesamiento
    # =========================
    def _show_external_data_info(self):
        """Mostrar información de los datos externos en un modal"""
        top = tk.Toplevel(self)
        top.title("Información de Datos Externos")
        top.geometry("500x400")
        top.grab_set()
        top.transient(self)
        
        frame = ttk.Frame(top, padding=20)
        frame.pack(fill=BOTH, expand=True)
        
        ttk.Label(frame, text="Información de Base de Datos", font=("Segoe UI", 16, "bold")).pack(pady=(0, 20))
        
        # Obtener información de la base de datos
        try:
            from database_manager import get_all_records_as_dataframe
            df = get_all_records_as_dataframe()
            
            if df.empty:
                total_records = 0
                date_range = "No disponible"
                last_import = "No disponible"
                unique_services = 0
                malignant_count = 0
            else:
                total_records = len(df)
                date_range = "Disponible" 
                last_import = "Disponible"
                unique_services = df.get('Servicio', pd.Series()).nunique() if 'Servicio' in df.columns else 0
                malignant_count = (df.get('Malignidad', pd.Series()).str.contains('PRESENTE', case=False, na=False)).sum() if 'Malignidad' in df.columns else 0
            
            info_text = f"""Total de informes en BD: {total_records}

Rango de fechas: {date_range}

Última importación: {last_import}

Servicios únicos: {unique_services}

Informes con malignidad: {malignant_count}"""
            
        except Exception as e:
            info_text = f"Error al obtener información de la base de datos:\n{str(e)}"
        
        text_widget = tk.Text(frame, wrap="word", font=("Segoe UI", 11))
        text_widget.pack(fill=BOTH, expand=True, pady=(0, 20))
        text_widget.insert("1.0", info_text)
        text_widget.configure(state="disabled")
        
        ttk.Button(frame, text="Cerrar", command=top.destroy, bootstyle="primary").pack()

    def _select_pdf_file(self):
        """Seleccionar un archivo PDF individual"""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo PDF",
            filetypes=[("Archivos PDF", "*.pdf")],
            initialdir=os.path.join(os.getcwd(), "pdfs_patologia") if os.path.exists("pdfs_patologia") else os.getcwd()
        )
        if file_path:
            try:
                records = self._process_file(file_path)
                messagebox.showinfo("Procesamiento", f"✅ Archivo procesado exitosamente:\n{records} registros extraídos")
                
                # Actualizar la vista de datos y estadísticas
                self.refresh_data_and_table()
                self._refresh_database_stats()
                self._refresh_files_list()
                
            except Exception as e:
                messagebox.showerror("Error", f"❌ Error procesando el archivo:\n{str(e)}")

    def _select_pdf_folder(self):
        """Seleccionar una carpeta con archivos PDF"""
        folder_path = filedialog.askdirectory(
            title="Seleccionar carpeta con PDFs",
            initialdir=os.path.join(os.getcwd(), "pdfs_patologia") if os.path.exists("pdfs_patologia") else os.getcwd()
        )
        if folder_path:
            pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]
            if pdf_files:
                for pdf_file in pdf_files:
                    self._process_file(os.path.join(folder_path, pdf_file))
                messagebox.showinfo("Procesamiento", f"Se procesaron {len(pdf_files)} archivos PDF.")
            else:
                messagebox.showwarning("Sin archivos", "No se encontraron archivos PDF en la carpeta seleccionada.")

    def _refresh_files_list(self):
        """Actualizar la lista de archivos en la carpeta pdfs_patologia"""
        pdfs_path = os.path.join(os.getcwd(), "pdfs_patologia")
        
        # Crear la carpeta si no existe
        if not os.path.exists(pdfs_path):
            os.makedirs(pdfs_path)
        
        # Limpiar la lista actual
        self.files_listbox.delete(0, tk.END)
        
        # Obtener archivos PDF
        try:
            pdf_files = [f for f in os.listdir(pdfs_path) if f.lower().endswith('.pdf')]
            pdf_files.sort()
            
            for pdf_file in pdf_files:
                self.files_listbox.insert(tk.END, pdf_file)
                
            if not pdf_files:
                self.files_listbox.insert(tk.END, "(No hay archivos PDF)")
                
        except Exception as e:
            self.files_listbox.insert(tk.END, f"Error: {str(e)}")

    def _process_selected_files(self):
        """Procesar los archivos seleccionados de la lista"""
        selected_indices = self.files_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("Sin selección", "Por favor seleccione uno o más archivos para procesar.")
            return
        
        pdfs_path = os.path.join(os.getcwd(), "pdfs_patologia")
        processed_count = 0
        total_records = 0
        errors = []
        
        # Procesar cada archivo seleccionado
        for index in selected_indices:
            filename = self.files_listbox.get(index)
            if filename != "(No hay archivos PDF)" and not filename.startswith("Error:"):
                file_path = os.path.join(pdfs_path, filename)
                if os.path.exists(file_path):
                    try:
                        records = self._process_file(file_path)
                        processed_count += 1
                        total_records += records
                    except Exception as e:
                        error_msg = f"{filename}: {str(e)}"
                        errors.append(error_msg)
                        print(f"❌ Error procesando {filename}: {e}")
        
        # Mostrar resultado final
        if processed_count > 0:
            success_msg = f"✅ Procesamiento completado:\n"
            success_msg += f"• {processed_count} archivos procesados\n"
            success_msg += f"• {total_records} registros extraídos y guardados en BD"
            
            if errors:
                success_msg += f"\n\n⚠️ Errores en {len(errors)} archivos:\n"
                success_msg += "\n".join(errors[:3])  # Mostrar solo los primeros 3 errores
                if len(errors) > 3:
                    success_msg += f"\n... y {len(errors) - 3} errores más."
            
            messagebox.showinfo("Procesamiento completado", success_msg)
            
            # Actualizar la vista de datos y estadísticas
            try:
                self.refresh_data_and_table()
                self._refresh_database_stats()  # Actualizar estadísticas de la BD
            except Exception as e:
                print(f"Error actualizando vista: {e}")
                
        else:
            error_msg = "❌ No se pudo procesar ningún archivo.\n\nErrores encontrados:\n"
            error_msg += "\n".join(errors[:5])  # Mostrar los primeros 5 errores
            messagebox.showerror("Error de procesamiento", error_msg)

    def _process_file(self, file_path):
        """Procesar un archivo PDF individual"""
        try:
            filename = os.path.basename(file_path)
            self.set_status(f"Procesando {filename}...")
            
            # Determinar el tipo de archivo y procesarlo adecuadamente
            if self._is_ihq_file(filename, file_path):
                # Procesar como archivo IHQ (biomarcadores)
                records_processed = self._process_ihq_file(file_path)
                self.set_status(f"✅ {filename}: {records_processed} registros IHQ procesados")
            else:
                # Procesar como archivo general de patología
                records_processed = self._process_general_file(file_path)
                self.set_status(f"✅ {filename}: {records_processed} registros generales procesados")
            
            print(f"✅ Procesamiento completado: {file_path} - {records_processed} registros")
            return records_processed
            
        except Exception as e:
            error_msg = f"Error procesando {filename}: {str(e)}"
            self.set_status(f"❌ {error_msg}")
            raise Exception(error_msg)

    def _is_ihq_file(self, filename, file_path):
        """Determinar si un archivo es de IHQ basándose en el nombre y contenido"""
        # Criterio 1: Nombre del archivo
        if "ihq" in filename.lower():
            return True
        
        # Criterio 2: Revisar contenido del archivo (muestra)
        try:
            from ocr_processing import pdf_to_text_enhanced
            # Solo leer una página para determinar el tipo
            import fitz
            doc = fitz.open(file_path)
            if len(doc) > 0:
                page_text = doc[0].get_text()
                doc.close()
                # Buscar indicadores de IHQ
                ihq_indicators = ['ihq', 'inmunohistoquimica', 'her2', 'ki-67', 'receptor estrogeno']
                return any(indicator in page_text.lower() for indicator in ihq_indicators)
        except Exception:
            pass
        
        return False

    def _process_ihq_file(self, file_path):
        """Procesar archivo IHQ usando el procesador especializado"""
        try:
            from procesador_ihq_biomarcadores import process_ihq_paths
            
            # Crear directorio de salida temporal si no existe
            output_dir = os.path.join(os.getcwd(), "EXCEL")
            os.makedirs(output_dir, exist_ok=True)
            
            # Procesar el archivo IHQ
            records_count = process_ihq_paths([file_path], output_dir)
            return records_count
            
        except Exception as e:
            raise Exception(f"Error en procesamiento IHQ: {str(e)}")

    def _process_general_file(self, file_path):
        """Procesar archivo general usando el procesador estándar"""
        try:
            # Importar los módulos necesarios
            from ocr_processing import pdf_to_text_enhanced
            import procesador_ihq as ihq
            import database_manager
            
            # Extraer texto del PDF
            full_text = pdf_to_text_enhanced(file_path)
            if not isinstance(full_text, str):
                full_text = '\n'.join(full_text)
            
            # Segmentar por informes (usando lógica similar a IHQ pero más general)
            records = []
            
            # Para archivos generales, puede haber múltiples informes
            # Intentar segmentar por "N. petición" o números de orden
            segments = self._segment_general_reports(full_text)
            
            if not segments:
                # Si no se puede segmentar, procesar como un solo informe
                segments = [full_text]
            
            for segment in segments:
                # Extraer datos base usando el procesador IHQ (que maneja datos generales también)
                base_data = ihq.extract_ihq_data(segment)
                base_rows = ihq.map_to_excel_format(base_data)
                
                if base_rows:
                    records.extend(base_rows)
            
            if not records:
                raise RuntimeError("No se pudo extraer información del PDF.")
            
            # Guardar en base de datos
            from database_manager import init_db, save_records
            init_db()
            save_records(records)
            
            return len(records)
            
        except Exception as e:
            raise Exception(f"Error en procesamiento general: {str(e)}")

    def _segment_general_reports(self, text):
        """Segmentar texto en informes individuales para archivos generales"""
        segments = []
        
        # Buscar patrones de número de petición más generales
        patterns = [
            r'(?i)(?:N[°.\s]*|No\.\s*|Nº\s*|N\s*)?petici[oó]n\s*[:\-]?\s*([A-Z]?\d{6,})',
            r'(?i)(?:registro|orden|numero|no\.?)\s*[:\-]?\s*(\d{4,})',
            r'(?i)(\d{4,})\s*(?:patolog[ií]a|biopsia|citolog[ií]a)'
        ]
        
        # Intentar segmentar por cualquiera de los patrones
        for pattern in patterns:
            matches = list(re.finditer(pattern, text))
            if len(matches) > 1:  # Solo si encontramos múltiples coincidencias
                starts = [(m.start(), m.group(1)) for m in matches]
                starts.sort()
                
                for i, (start, code) in enumerate(starts):
                    if i + 1 < len(starts):
                        end = starts[i + 1][0]
                        segment = text[start:end].strip()
                    else:
                        segment = text[start:].strip()
                    
                    if len(segment) > 100:  # Solo segmentos con contenido suficiente
                        segments.append(segment)
                
                break  # Usar el primer patrón que funcione
        
        return segments

    def _export_selected_data(self):
        """Exportar todos los datos completos de la base de datos a Excel con presentación profesional"""
        try:
            # Obtener todos los datos de la base de datos
            from database_manager import get_all_records_as_dataframe
            df_complete = get_all_records_as_dataframe()
            
            if df_complete.empty:
                messagebox.showwarning("Advertencia", "No hay datos en la base de datos para exportar.")
                return
                
            # Diálogo para guardar archivo
            file_path = filedialog.asksaveasfilename(
                title="Exportar Base de Datos Completa - EVARISIS",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if file_path:
                # Crear el archivo Excel con múltiples hojas
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    
                    # ========== HOJA 1: PRESENTACIÓN ==========
                    # Crear DataFrame para la presentación
                    presentation_data = [
                        ["EVARISIS Gestor HUV - Oncología", ""],
                        ["Exportación de Base de Datos", ""],
                        ["", ""],
                        ["Información del Reporte", ""],
                        ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
                        ["Fecha y Hora de Exportación:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
                        ["Usuario:", self.info_usuario.get("nombre", "Sistema")],
                        ["Cargo:", self.info_usuario.get("cargo", "N/A")],
                        ["", ""],
                        ["Estadísticas de la Base de Datos", ""],
                        ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
                        ["Total de Registros:", len(df_complete)],
                        ["Número de Campos:", len(df_complete.columns)],
                        ["", ""],
                    ]
                    
                    # Calcular estadísticas adicionales
                    fecha_cols = [col for col in df_complete.columns if 'fecha' in col.lower()]
                    if fecha_cols:
                        try:
                            for col in fecha_cols:
                                fechas = pd.to_datetime(df_complete[col], errors='coerce')
                                fechas_validas = fechas.dropna()
                                if not fechas_validas.empty:
                                    fecha_min = fechas_validas.min().strftime("%d/%m/%Y")
                                    fecha_max = fechas_validas.max().strftime("%d/%m/%Y")
                                    presentation_data.extend([
                                        [f"Rango de Fechas ({col}):", f"{fecha_min} - {fecha_max}"],
                                    ])
                                    break
                        except:
                            pass
                    
                    # Servicios únicos
                    servicio_cols = [col for col in df_complete.columns if 'servicio' in col.lower()]
                    if servicio_cols:
                        unique_services = df_complete[servicio_cols[0]].nunique()
                        presentation_data.append(["Servicios Únicos:", unique_services])
                    
                    # Casos malignos
                    malignidad_cols = [col for col in df_complete.columns if 'malign' in col.lower()]
                    if malignidad_cols:
                        malignant_count = (df_complete[malignidad_cols[0]].str.contains('PRESENTE', case=False, na=False)).sum()
                        presentation_data.append(["Casos con Malignidad:", malignant_count])
                    
                    presentation_data.extend([
                        ["", ""],
                        ["Descripción", ""],
                        ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
                        ["Este archivo contiene la exportación completa", ""],
                        ["de la base de datos del sistema EVARISIS", ""],
                        ["Gestor HUV - Oncología.", ""],
                        ["", ""],
                        ["La información incluye todos los campos", ""],
                        ["almacenados para cada informe médico:", ""],
                        ["- Datos del paciente", ""],
                        ["- Información clínica", ""],
                        ["- Resultados de laboratorio", ""],
                        ["- Análisis histopatológicos", ""],
                        ["- Biomarcadores (IHQ)", ""],
                        ["- Fechas y responsables", ""],
                        ["", ""],
                        ["Hospital Universitario del Valle", ""],
                        ["Sistema EVARISIS © 2025", ""],
                    ])
                    
                    df_presentation = pd.DataFrame(presentation_data, columns=["Campo", "Valor"])
                    df_presentation.to_excel(writer, sheet_name='Presentación', index=False)
                    
                    # ========== HOJA 2: DATOS COMPLETOS ==========
                    df_complete.to_excel(writer, sheet_name='Datos_Completos', index=False)
                    
                    # ========== FORMATO ==========
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    
                    # Formatear hoja de presentación
                    ws_pres = writer.sheets['Presentación']
                    
                    # Título principal
                    ws_pres['A1'].font = Font(size=18, bold=True, color="FFFFFF")
                    ws_pres['A1'].fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
                    ws_pres['A1'].alignment = Alignment(horizontal="center", vertical="center")
                    ws_pres.merge_cells('A1:B1')
                    
                    # Subtítulo
                    ws_pres['A2'].font = Font(size=14, bold=True, color="FFFFFF")
                    ws_pres['A2'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                    ws_pres['A2'].alignment = Alignment(horizontal="center", vertical="center")
                    ws_pres.merge_cells('A2:B2')
                    
                    # Secciones
                    section_font = Font(size=12, bold=True, color="1B4F72")
                    for row in range(1, len(presentation_data) + 2):
                        cell_value = ws_pres[f'A{row}'].value
                        if cell_value and ("Información" in str(cell_value) or "Estadísticas" in str(cell_value) or "Descripción" in str(cell_value)):
                            ws_pres[f'A{row}'].font = section_font
                            ws_pres.merge_cells(f'A{row}:B{row}')
                    
                    # Ajustar ancho de columnas
                    ws_pres.column_dimensions['A'].width = 40
                    ws_pres.column_dimensions['B'].width = 30
                    
                    # Formatear hoja de datos
                    ws_data = writer.sheets['Datos_Completos']
                    
                    # Headers
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    for cell in ws_data[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Ajustar ancho de columnas automáticamente
                    for column in ws_data.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                        ws_data.column_dimensions[column_letter].width = adjusted_width
                
                messagebox.showinfo(
                    "Exportación Exitosa", 
                    f"✅ Base de datos exportada exitosamente!\n\n"
                    f"📊 {len(df_complete)} registros exportados\n"
                    f"📋 {len(df_complete.columns)} campos por registro\n"
                    f"📁 Archivo: {os.path.basename(file_path)}\n\n"
                    f"El archivo incluye:\n"
                    f"• Hoja de presentación con estadísticas\n"
                    f"• Datos completos de todos los registros"
                )
                
        except Exception as e:
            messagebox.showerror("Error", f"❌ Error al exportar:\n{str(e)}")
            print(f"Error detallado: {e}")

    def _export_current_record(self):
        """Exportar el registro actualmente seleccionado"""
        try:
            selection = self.tree.selection()
            if not selection:
                messagebox.showwarning("Advertencia", "No hay registro seleccionado.")
                return
                
            # Tomar solo el primer registro seleccionado
            item = selection[0]
            self._export_selected_data()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar registro: {str(e)}")

    def _copy_details(self):
        """Copiar detalles del registro al clipboard"""
        try:
            details_text = self.detail_textbox.get("1.0", tk.END)
            if details_text.strip():
                self.clipboard_clear()
                self.clipboard_append(details_text)
                messagebox.showinfo("Copiado", "Detalles copiados al portapapeles")
            else:
                messagebox.showwarning("Advertencia", "No hay detalles para copiar")
        except Exception as e:
            messagebox.showerror("Error", f"Error al copiar: {str(e)}")

    def _on_double_click(self, event):
        """Manejar doble clic en la tabla"""
        selection = self.tree.selection()
        if selection:
            # Expandir detalles o realizar acción específica
            self.mostrar_detalle_registro(event)

    def _init_treeview_style(self):
        """
        Define el estilo 'Custom.Treeview' usando los colores del tema actual de TTKBootstrap.
        """
        try:
            from tkinter import ttk as _ttk
            s = _ttk.Style()

            # Usar colores del tema TTKBootstrap actual
            bg_color = self.style.colors.bg or "#ffffff"
            fg_color = self.style.colors.fg or "#000000"  
            primary_color = self.style.colors.primary or "#0d6efd"
            secondary_color = self.style.colors.secondary or "#6c757d"

            s.configure(
                "Custom.Treeview",
                background=bg_color,
                fieldbackground=bg_color,
                foreground=fg_color,
                rowheight=26,
                borderwidth=0,
            )
            s.map(
                "Custom.Treeview",
                background=[("selected", primary_color)],
                foreground=[("selected", "white")],
            )
            s.configure(
                "Custom.Treeview.Heading",
                background=secondary_color,
                foreground="white",
                relief="flat",
                padding=6,
            )
        except Exception as e:
            print(f"[WARN] No se pudo configurar Custom.Treeview: {e}")

def main():
    """
    Función principal que configura el entorno y lanza la aplicación.
    """
    print("Iniciando EVARISIS Gestor H.U.V...")
    
    # Configurar el parser de argumentos
    parser = argparse.ArgumentParser(description="EVARISIS Gestor HUV - Oncología")
    
    parser.add_argument("--nombre", type=str, default="Usuario Sistema", help="Nombre del usuario logueado.")
    parser.add_argument("--cargo", type=str, default="Administrador", help="Cargo del usuario logueado.")
    parser.add_argument("--foto", type=str, default="SIN_FOTO", help="Ruta a la foto de perfil del usuario.")
    parser.add_argument("--tema", type=str, default="dark", help="Tema visual de la aplicación.")
    parser.add_argument("--ruta-fotos", type=str, default="", help="Ruta al directorio base de fotos de usuarios.")
    parser.add_argument(
        "--lanzado-por-evarisis",
        action='store_true',
        help="Bandera interna para verificar el lanzamiento desde la app principal."
    )
    parser.add_argument(
        "--modo-independiente",
        action='store_true',
        help="Permite ejecutar la aplicación de forma independiente (sin EVARISIS)."
    )
    
    args = parser.parse_args()

    # Verificación de seguridad (permite modo independiente)
    if not args.lanzado_por_evarisis and not args.modo_independiente:
        import tkinter as tk
        root_error = tk.Tk()
        root_error.withdraw()
        messagebox.showerror(
            "Acceso no autorizado",
            "Esta aplicación no puede ejecutarse directamente.\n\n"
            "Para uso independiente, use: python ui.py --modo-independiente\n"
            "Para integración EVARISIS, use la bandera --lanzado-por-evarisis"
        )
        sys.exit("Ejecución no autorizada. Terminando proceso.")

    # 1. Configuramos Tesseract antes de que cualquier otra cosa lo necesite
    configure_tesseract()

    # Configurar información del usuario
    info_usuario_recibida = {
        "nombre": args.nombre,
        "cargo": args.cargo,
        "ruta_foto": args.foto,
        "ruta_directorio_fotos": args.ruta_fotos
    }
    
    # Configurar logging
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    # Mapear tema del argumento a tema TTKBootstrap
    tema_ttk = THEME_MAP.get(args.tema, "darkly")
    
    # Crear y ejecutar la aplicación
    app = App(info_usuario=info_usuario_recibida, tema=tema_ttk)
    app.mainloop()
    
    print("Aplicación cerrada. ¡Hasta luego!")

if __name__ == "__main__":
    main()